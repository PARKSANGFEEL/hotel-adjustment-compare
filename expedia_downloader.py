# -*- coding: utf-8 -*-
"""
Expedia 명세서 자동 다운로드 모듈


기능:
1. Expedia Partner Central 로그인 (쿠키 기반)
2. Statements & Invoices 페이지 접근
3. 명세서 목록 조회
4. 명세서 파일 다운로드 (TXT 형식)

작성자: GitHub Copilot
날짜: 2026-01-02
"""

import os
import time
import json
import urllib.request
from pathlib import Path
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager


class ExpediaDownloader:
    """
    Expedia Partner Central에서 명세서를 자동으로 다운로드하는 클래스
    """
    
    def __init__(self, username=None, password=None, download_dir='ota-adjustment', base_dir=None):
        """
        초기화
        
        Args:
            username: Expedia 로그인 아이디 (env에서 가져옴)
            password: Expedia 로그인 비밀번호 (env에서 가져옴)
            download_dir: 다운로드 저장 폴더
            base_dir: 작업 디렉토리 (기본값: 현재 디렉토리)
        """
        self.username = username or os.environ.get('EXPEDIA_USERNAME')
        self.password = password or os.environ.get('EXPEDIA_PASSWORD')

        # 경로 설정
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.download_dir = self.base_dir / download_dir
        self.download_dir.mkdir(exist_ok=True)

        # 쿠키 저장 파일
        self.cookies_file = self.base_dir / 'expedia_cookies.json'

        # 크리덴셜 없이 쿠키 로그인만 시도할 수 있게 허용
        if (not self.username or not self.password) and not self.cookies_file.exists():
            raise ValueError("EXPEDIA_USERNAME 및 EXPEDIA_PASSWORD 환경변수가 필요합니다")
        if (not self.username or not self.password) and self.cookies_file.exists():
            print("[INFO] 환경변수 없이 저장된 쿠키로만 로그인 시도")
        
        # 브라우저
        self.driver = None
        
    def setup_driver(self):
        """Chrome 드라이버 설정 (샌드박스 비활성화)"""
        chrome_options = Options()
        
        # Windows 환경에서 필요한 옵션
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # 다운로드 설정
        temp_download_dir = str(self.base_dir / 'temp_downloads')
        Path(temp_download_dir).mkdir(exist_ok=True)
        
        prefs = {
            'download.default_directory': temp_download_dir,
            'download.prompt_for_download': False,
            'download.directory_upgrade': True,
            'safebrowsing.enabled': True,
            'profile.default_content_setting_values.automatic_downloads': 1  # 다중 다운로드 자동 허용
        }
        chrome_options.add_experimental_option('prefs', prefs)
        
        # ChromeDriver 자동 설치 및 설정
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            print("Chrome 드라이버 자동 설치 및 설정 완료")
        except Exception as e:
            print(f"ChromeDriver 자동 설치 실패, 시스템 드라이버 사용 시도: {e}")
            self.driver = webdriver.Chrome(options=chrome_options)
            print("시스템 Chrome 드라이버 사용")
        
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        self.temp_download_dir = temp_download_dir
        
        print("Chrome 드라이버 설정 완료")
        
    def load_cookies(self):
        """저장된 쿠키 로드"""
        if not self.cookies_file.exists():
            print("[쿠키] 저장된 쿠키 없음")
            return False

        try:
            with open(self.cookies_file, 'r') as f:
                cookies = json.load(f)

            # 쿠키를 추가하기 전에 도메인 페이지 한번 열기
            self.driver.get("https://www.expediapartnercentral.com/Account/Logon?signedOff=true")
            time.sleep(2)

            for cookie in cookies:
                try:
                    self.driver.add_cookie(cookie)
                except Exception as e:
                    print(f"  [경고] 쿠키 추가 실패 ({cookie.get('name')}): {e}")
            
            print(f"  {len(cookies)}/{len(cookies)} 쿠키 로드 완료")
            return True
            
        except Exception as e:
            print(f"[ERROR] 쿠키 로드 실패: {e}")
            return False

    def login(self):
        """쿠키 우선 로그인 (환경변수 없으면 쿠키만 시도)"""
        login_url = "https://www.expediapartnercentral.com/Account/Logon?signedOff=true"
        statements_url = "https://apps.expediapartnercentral.com/lodging/accounting/statementsAndInvoices.html?htid=17300293&tab=invoices"

        print(f"\n[로그인] {login_url}")
        try:
            # 로그인 페이지 접속
            self.driver.get(login_url)
            time.sleep(2)

            # 쿠키 로드 시도
            loaded = self.load_cookies()

            if loaded:
                print("\n[시도] 저장된 쿠키로 로그인...")
                self.driver.get(statements_url)
                time.sleep(5)
                current_url = self.driver.current_url
                if 'accounting/statementsAndInvoices' in current_url:
                    # 페이로드 존재 확인
                    for _ in range(10):
                        if 'statementsAndInvoicesPayload' in self.driver.page_source:
                            print("  [SUCCESS] 저장된 쿠키로 로그인 성공")
                            return True
                        time.sleep(1)
                print("  [INFO] 쿠키 로그인 실패 또는 추가 인증 필요")

            # 쿠키만 있고 크리덴셜 없으면 종료
            if (not self.username or not self.password):
                print("  [ERROR] 로그인 실패 (쿠키 인증 실패, 자격 증명 없음)")
                return False

            # 수동 로그인 미구현
            print("  [ERROR] 로그인 실패 (수동 로그인 미구현)")
            return False

        except Exception as e:
            print(f"  [ERROR] 로그인 오류: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def save_cookies(self):
        """현재 쿠키 저장"""
        try:
            cookies = self.driver.get_cookies()
            with open(self.cookies_file, 'w') as f:
                json.dump(cookies, f, indent=2)
            print(f"[쿠키] {len(cookies)}개 쿠키 저장됨")
            return True
        except Exception as e:
            print(f"[ERROR] 쿠키 저장 실패: {e}")
            return False
    
        script = f"""
        const target = '{invoice_id}'.replace(/\\D/g, '');
        const debugTexts = [];
        const rows = document.querySelectorAll('.fds-table tbody tr');

        const collectDebug = (el) => {{
            if (debugTexts.length < 15) {{ debugTexts.push((el.textContent || '').trim()); }}
        }};

        for (const row of rows) {{
            const btn = row.querySelector('button');
            if (btn) {{
                const digits = (btn.textContent || '').replace(/\\D/g, '');
                collectDebug(btn);
                if (digits === target) {{ btn.click(); return {{ 'found': true, 'debugTexts': debugTexts }}; }}
            }}

            const anchors = row.querySelectorAll('a');
            for (const a of anchors) {{
                const digits = (a.textContent || '').replace(/\\D/g, '');
                collectDebug(a);
                if (digits === target) {{ a.click(); return {{ 'found': true, 'debugTexts': debugTexts }}; }}
            }}

            const spans = row.querySelectorAll('td[data-field-label], span, div');
            for (const s of spans) {{
                const digits = (s.textContent || '').replace(/\\D/g, '');
                collectDebug(s);
                if (digits === target) {{
                    const clickable = s.closest('a, button') || s;
                    clickable.click();
                    return {{ 'found': true, 'debugTexts': debugTexts }};
                }}
            }}
        }}

        const allClickables = document.querySelectorAll('button, a');
        for (const el of allClickables) {{
            const digits = (el.textContent || '').replace(/\\D/g, '');
            collectDebug(el);
            if (digits === target) {{ el.click(); return {{ 'found': true, 'debugTexts': debugTexts }}; }}
        }}

        return {{ 'found': false, 'debugTexts': debugTexts }};
        """
        """수동 로그인 (2단계)"""
        print("\n[수동 로그인] 시작...")
        
        try:
            # 1단계: 이메일 입력
            print("\n[1단계] 이메일 입력")
            time.sleep(3)
            
            # 스크린샷 저장
            screenshot_path = self.base_dir / 'login_step1.png'
            self.driver.save_screenshot(str(screenshot_path))
            print(f"  스크린샷: {screenshot_path}")
            
            # 이메일 입력 필드 찾기 (id="emailControl")
            try:
                email_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'emailControl'))
                )
                print(f"  [OK] 이메일 필드 발견")
            except:
                # 대체 선택자 시도
                email_field = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="username"][type="email"]'))
                )
                print(f"  [OK] 이메일 필드 발견 (대체 선택자)")
            
            email_field.clear()
            email_field.send_keys(self.username)
            print(f"  이메일 입력 완료: {self.username}")
            time.sleep(1)
            
            # 다음 버튼 클릭 (id="continueButton")
            try:
                continue_button = self.driver.find_element(By.ID, 'continueButton')
                print(f"  [OK] 다음 버튼 발견")
            except:
                # 대체 선택자
                continue_button = self.driver.find_element(By.CSS_SELECTOR, 'button[type="submit"].continue-button')
                print(f"  [OK] 다음 버튼 발견 (대체 선택자)")
            
            continue_button.click()
            print(f"  다음 버튼 클릭")
            
            # 2단계: 비밀번호 입력
            print("\n[2단계] 비밀번호 입력")
            time.sleep(5)  # 새 페이지 로딩 대기
            
            # 스크린샷 저장
            screenshot_path = self.base_dir / 'login_step2.png'
            self.driver.save_screenshot(str(screenshot_path))
            print(f"  스크린샷: {screenshot_path}")
            
            # 비밀번호 입력 필드 찾기 (id="password-input")
            try:
                password_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'password-input'))
                )
                print(f"  [OK] 비밀번호 필드 발견")
            except:
                # 대체 선택자
                password_field = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="password"][type="password"]'))
                )
                print(f"  [OK] 비밀번호 필드 발견 (대체 선택자)")
            
            password_field.clear()
            password_field.send_keys(self.password)
            print(f"  비밀번호 입력 완료")
            time.sleep(1)
            
            # 계속하기 버튼 클릭 (id="password-continue")
            try:
                submit_button = self.driver.find_element(By.ID, 'password-continue')
                print(f"  [OK] 계속하기 버튼 발견")
            except:
                # 대체 선택자
                submit_button = self.driver.find_element(By.CSS_SELECTOR, 'button[type="submit"].uitk-button-primary')
                print(f"  [OK] 계속하기 버튼 발견 (대체 선택자)")
            
            submit_button.click()
            print(f"  계속하기 버튼 클릭")
            
            # 2차 인증 처리 (필요한 경우)
            print("\n[2차 인증] 확인 중...")
            time.sleep(10)
            
            # 로그인 후 스크린샷
            after_screenshot = self.base_dir / 'after_login.png'
            self.driver.save_screenshot(str(after_screenshot))
            print(f"  로그인 후 스크린샷: {after_screenshot}")
            
            current_url = self.driver.current_url
            print(f"  현재 URL: {current_url}")
            
            # 로그인 성공 확인
            if 'Logon' not in current_url and 'login' not in current_url.lower():
                print("  [SUCCESS] 로그인 성공")
                self.save_cookies()  # 쿠키 저장
                return True
            else:
                print("  [FAILED] 로그인 실패 (여전히 로그인 페이지)")
                return False
                
        except Exception as e:
            print(f"  [ERROR] 로그인 오류: {e}")
            import traceback
            traceback.print_exc()
            
            # 에러 시 스크린샷
            error_screenshot = self.base_dir / 'error_screenshot.png'
            try:
                self.driver.save_screenshot(str(error_screenshot))
                print(f"  에러 스크린샷: {error_screenshot}")
            except:
                pass
            
            return False
    
    def navigate_to_statements(self):
        """명세서 페이지로 이동"""
        statements_url = "https://apps.expediapartnercentral.com/lodging/accounting/statementsAndInvoices.html?htid=17300293&tab=invoices"
        
        print("\n[이동] 명세서 페이지")
        
        try:
            self.driver.get(statements_url)
            time.sleep(5)
            print(f"  현재 URL: {self.driver.current_url}")
            
            # statements 탭 클릭
            try:
                statements_tab = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '[data-content-id="statements"]'))
                )
                print("  명세서 탭 발견, 클릭 중...")
                statements_tab.click()
                time.sleep(5)
            except:
                print("  [INFO] 명세서 탭을 찾을 수 없음 (이미 선택되었을 수 있음)")
            
            # JavaScript 페이로드 로드 대기 (최대 30초)
            print("  JavaScript 데이터 로드 대기 중...")
            for i in range(45):
                page_source = self.driver.page_source
                if 'statementsAndInvoicesPayload' in page_source:
                    print("  명세서 페이지 로드 완료")
                    break
                time.sleep(1)
            else:
                print("  [WARNING] JavaScript 데이터 로드 타임아웃 (계속 진행)")
            
            # 페이지당 표시 개수를 100개로 변경
            self._set_page_size_to_100()
            
            return True
            
        except Exception as e:
            print(f"  [ERROR] 명세서 페이지 이동 실패: {e}")
            return False
    
    def _set_page_size_to_100(self):
        """페이지당 표시 개수를 100개로 설정"""
        try:
            print("  [시도] 페이지 크기를 100으로 설정...")
            
            # === 전략 1: visible한 select 직접 변경 ===
            selects = self.driver.find_elements(By.CSS_SELECTOR, 'select.fds-field-select')
            print(f"  [INFO] {len(selects)}개의 select 발견")
            
            visible_select = None
            for idx, sel in enumerate(selects):
                parent_visible = self.driver.execute_script("return arguments[0].parentElement.offsetParent !== null", sel)
                if parent_visible:
                    visible_select = sel
                    print(f"  [OK] Visible select 발견 (index {idx})")
                    break
            
            if not visible_select:
                print("  [WARNING] Visible select를 찾지 못함 (대안 시도)")
            
            if visible_select:
                # JavaScript로 값 변경 + 이벤트 발생
                self.driver.execute_script("""
                    const select = arguments[0];
                    select.value = '100';
                    ['input', 'change', 'blur', 'click'].forEach(eventType => {
                        const event = new Event(eventType, { bubbles: true, cancelable: true });
                        select.dispatchEvent(event);
                    });
                    const changeEvent = new Event('change', { bubbles: true });
                    select.parentElement.dispatchEvent(changeEvent);
                """, visible_select)
                print("  [OK] select 값을 100으로 변경")
                time.sleep(1)
                new_value = visible_select.get_attribute('value')
                print(f"  [INFO] 변경 후 값: {new_value}")
                if new_value == '100':
                    print("  [SUCCESS] 페이지 크기가 100으로 변경됨!")
                else:
                    print(f"  [WARNING] 값이 {new_value}로 설정되었지만 100이 아님")
            
            # === 전략 2: 화면에 보이는 '10' 배지/버튼 클릭 시도 ===
            # (일부 UI에서 select는 숨겨져 있고, 상단의 badge를 눌러야 dropdown이 열림)
            try:
                badge = self.driver.execute_script("""
                    const candidates = Array.from(document.querySelectorAll('div, span, button'));
                    for (const el of candidates) {
                        if (!el.offsetParent) continue; // visible only
                        const text = (el.textContent || '').trim();
                        if (text === '10' || text === '10개') return el;
                    }
                    return null;
                """)
                if badge:
                    self.driver.execute_script("arguments[0].click();", badge)
                    time.sleep(0.5)
                    # dropdown이 열렸다고 가정하고 select 값을 다시 100으로 설정
                    if visible_select:
                        self.driver.execute_script("arguments[0].value='100'; arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", visible_select)
                        print("  [INFO] 배지 클릭 후 select 재설정 완료")
            except Exception:
                pass
            
            # === 테이블 로드 확인 ===
            try:
                for _ in range(10):
                    rows = self.driver.find_elements(By.CSS_SELECTOR, '.fds-table tbody tr, [role="table"] [role="row"]')
                    if len(rows) >= 50:
                        break
                    time.sleep(0.5)
                print(f"  [INFO] 테이블 행 개수: {len(rows)}")
            except Exception:
                pass
            
            return True
                
        except Exception as e:
            print(f"  [ERROR] 페이지 크기 변경 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_statement_list(self):
        """
        명세서 목록 조회 (JavaScript 페이로드에서 파싱)
        
        Returns:
            list: 명세서 정보 리스트
        """
        print("\n[조회] 명세서 목록")
        
        try:
            # 페이지 소스에서 JSON 데이터 추출
            page_source = self.driver.page_source
            
            # statementsAndInvoicesPayload 찾기 (JavaScript 변수로 정의됨)
            start_marker = 'statementsAndInvoicesPayload: '
            start_idx = page_source.find(start_marker)
            
            if start_idx == -1:
                print("  [ERROR] statementsAndInvoicesPayload를 찾을 수 없습니다")
                # 디버깅용 HTML 저장
                debug_path = self.base_dir / 'debug_no_payload.html'
                try:
                    debug_path.write_text(page_source, encoding='utf-8')
                    print(f"  [DEBUG] 페이지 소스를 저장했습니다: {debug_path}")
                except Exception as e:
                    print(f"  [DEBUG] 페이지 저장 실패: {e}")
                return []
            
            start_idx += len(start_marker)
            
            # JSON 객체 끝 찾기
            brace_count = 0
            in_string = False
            escape_next = False
            end_idx = start_idx
            
            for i in range(start_idx, len(page_source)):
                char = page_source[i]
                
                if escape_next:
                    escape_next = False
                    continue
                
                if char == '\\':
                    escape_next = True
                    continue
                
                if char == '"' and not escape_next:
                    in_string = not in_string
                    continue
                
                if not in_string:
                    if char == '{':
                        brace_count += 1
                    elif char == '}':
                        brace_count -= 1
                        if brace_count == 0:
                            end_idx = i + 1
                            break
            
            json_str = page_source[start_idx:end_idx]
            
            # JSON 파싱
            import json
            data = json.loads(json_str)
            
            statements = []
            
            # statements 데이터 추출
            if 'statements' in data and 'paymentList' in data['statements']:
                for stmt in data['statements']['paymentList']:
                    statement = {
                        'paymentRequestId': stmt.get('paymentRequestId', ''),
                        'invoiceId': stmt.get('invoiceId', ''),
                        'dateRequested': stmt.get('dateRequested', ''),
                        'paymentRequestStatus': stmt.get('paymentRequestStatus', ''),
                        'amountProcessed': stmt.get('amountProcessed', 0),
                        'amountProcessedCurrency': stmt.get('amountProcessedCurrency', ''),
                        'datePaid': stmt.get('datePaid', ''),
                        'paymentReferenceNumber': stmt.get('paymentReferenceNumber', ''),
                        'paymentRequestFilePath': stmt.get('paymentRequestFilePath', ''),
                        'paymentNoticePath': stmt.get('paymentNoticePath', '')
                    }
                    statements.append(statement)
                
                # datePaid가 있는 것만 필터링 (결제된 명세서만)
                paid_statements = [s for s in statements if s.get('datePaid')]
                print(f"  [OK] 총 {len(statements)}개 중 {len(paid_statements)}개 결제 명세서")
                return paid_statements
            
            # invoices 데이터 추출 (fallback)
            if 'invoices' in data and 'invoices' in data['invoices']:
                invoices = []
                for inv in data['invoices']['invoices']:
                    invoice = {
                        'transactionNumber': inv.get('transactionNumber', ''),
                        'transactionDate': inv.get('transactionDate', ''),
                        'transactionType': inv.get('transactionType', ''),
                        'originalAmount': inv.get('originalAmount', 0),
                        'transactionCurrency': inv.get('transactionCurrency', ''),
                        'status': inv.get('status', ''),
                        'pdfFilePath': inv.get('pdfFilePath', '')
                    }
                    invoices.append(invoice)
                
                print(f"  [OK] {len(invoices)}개 인보이스 발견")
                return invoices
            
            print("  [ERROR] statements 또는 invoices 데이터를 찾을 수 없습니다")
            return []
            
        except json.JSONDecodeError as e:
            print(f"  [ERROR] JSON 파싱 오류: {e}")
            print(f"  JSON 샘플: {json_str[:200] if 'json_str' in locals() else 'N/A'}")
            return []
        except Exception as e:
            print(f"  [ERROR] 명세서 목록 조회 오류: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _download_file_via_http(self, url, target_path):
        """Chrome 쿠키를 활용해 직접 파일을 내려받는다."""
        try:
            cookies = self.driver.get_cookies() if self.driver else []
            cookie_header = '; '.join([f"{c['name']}={c['value']}" for c in cookies if c.get('value')])
            if not cookie_header:
                print("  [INFO] 직접 다운로드용 쿠키가 없어 UI로 진행")
                return False

            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Cookie': cookie_header
            }
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=30) as resp:
                if resp.status != 200:
                    print(f"  [WARNING] 직접 다운로드 실패 (HTTP {resp.status})")
                    return False
                data = resp.read()

            if not data:
                print("  [WARNING] 직접 다운로드 응답이 비어 있음")
                return False

            target_path.write_bytes(data)
            print(f"  [OK] 직접 다운로드 완료: {target_path.name}")
            return True
        except Exception as e:
            print(f"  [INFO] 직접 다운로드 실패, UI 시도로 전환: {e}")
            return False

    def _click_invoice_button_on_current_page(self, invoice_id):
        """현재 페이지에서 인보이스 ID(지불 ID) 버튼/링크를 클릭한다."""
        script = f"""
        const target = '{invoice_id}'.replace(/\\D/g, '');
        const debugTexts = [];
        const rows = document.querySelectorAll('.fds-table tbody tr, [role="table"] [role="row"]:not([role="rowheader"])');

        const collectDebug = (el) => {{
            if (debugTexts.length < 15) {{ debugTexts.push((el.textContent || '').substring(0, 80)); }}
        }};

        // 테이블에 visible row가 적으면 경고
        if (rows.length < 3) {{
            console.log(`[WARN] Visible rows: ${{rows.length}} (maybe page size not increased)`);
        }}

        for (const row of rows) {{
            const btn = row.querySelector('button');
            if (btn) {{
                const digits = (btn.textContent || '').replace(/\\D/g, '');
                collectDebug(btn);
                if (digits === target) {{ btn.click(); return {{ 'found': true, 'debugTexts': debugTexts, 'method': 'button' }}; }}
            }}

            const anchors = row.querySelectorAll('a');
            for (const a of anchors) {{
                const digits = (a.textContent || '').replace(/\\D/g, '');
                collectDebug(a);
                if (digits === target) {{ a.click(); return {{ 'found': true, 'debugTexts': debugTexts, 'method': 'anchor' }}; }}
            }}

            const spans = row.querySelectorAll('td[data-field-label], span, div');
            for (const s of spans) {{
                const digits = (s.textContent || '').replace(/\\D/g, '');
                collectDebug(s);
                if (digits === target) {{
                    const clickable = s.closest('a, button, [role="button"]') || s;
                    clickable.click();
                    return {{ 'found': true, 'debugTexts': debugTexts, 'method': 'span' }};
                }}
            }}
        }}

        const allClickables = document.querySelectorAll('button, a, [role="button"]');
        for (const el of allClickables) {{
            const digits = (el.textContent || '').replace(/\\D/g, '');
            collectDebug(el);
            if (digits === target) {{ el.click(); return {{ 'found': true, 'debugTexts': debugTexts, 'method': 'global' }}; }}
        }}

        return {{ 'found': false, 'debugTexts': debugTexts, 'rowCount': rows.length, 'method': 'none' }};
        """
        try:
            result = self.driver.execute_script(script)
            if isinstance(result, dict):
                if result.get('found'):
                    print(f"  [OK] 인보이스 ID 찾음 (방법: {result.get('method')})")
                    return True
                # 테이블이 너무 작으면 경고
                row_count = result.get('rowCount', 0)
                if row_count < 5:
                    print(f"  [WARN] 테이블 행이 적음: {row_count}개 (pagination 필요)")
                dbg = result.get('debugTexts', [])
                if dbg:
                    print(f"  [DEBUG] 화면의 첫 15개 클릭텍스트: {dbg[:5]}")  # 처음 5개만 표시
            return False
        except Exception as e:
            print(f"  [DEBUG] 버튼 스크립트 실행 오류: {e}")
            return False

    def _go_to_next_page(self):
        """다음 페이지 화살표 클릭 시도"""
        try:
            btn = self.driver.execute_script("""
                const icon = document.querySelector('button .fds-icon-name-arrow-forward-ios');
                if (icon) {
                    const b = icon.closest('button');
                    if (b && !b.disabled && b.getAttribute('aria-disabled') !== 'true') {
                        b.click();
                        return true;
                    }
                }
                const candidates = Array.from(document.querySelectorAll('button, a'));
                for (const el of candidates) {
                    const disabled = el.disabled || el.getAttribute('aria-disabled') === 'true';
                    if (disabled) continue;
                    const hasIcon = el.querySelector('.fds-icon-name-arrow-forward-ios, .fds-icon-name-chevron-right, .fds-icon-name-arrow-forward, .fds-icon-name-arrow-right');
                    const hasUse = el.querySelector('use[href*="arrow-forward"], use[xlink\:href*="arrow-forward"], use[href*="chevron-right"], use[xlink\:href*="chevron-right"]');
                    const aria = (el.getAttribute('aria-label') || '').toLowerCase();
                    if (hasIcon || hasUse || aria.includes('next') || aria.includes('다음')) {
                        el.click();
                        return true;
                    }
                }
                return false;
            """)
            if btn:
                print("  [INFO] 다음 페이지로 이동")
                time.sleep(0.8)
                return True
            print("  [INFO] 다음 페이지 버튼 없음 또는 비활성")
            return False
        except Exception as e:
            print(f"  [DEBUG] 다음 페이지 이동 실패: {e}")
            return False

    def _wait_for_detail_page(self, prev_url):
        """지불 ID 클릭 후 상세 화면이 로드될 때까지 대기한다."""
        for _ in range(30):  # 최대 30초
            cur_url = self.driver.current_url
            page = self.driver.page_source
            if cur_url != prev_url:
                return True
            if 'PSLCreate' in page:
                return True
            if 'statementsAndInvoicesPayload' not in page:
                return True
            time.sleep(1)
        return False

    def _click_table_download_icon(self):
        """예약 테이블의 다운로드 아이콘을 클릭한다 (CSV/파일 다운로드)."""
        try:
            icon = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".statements-and-invoices-download-ecInvoices-icon"))
            )
            anchor = self.driver.execute_script("return arguments[0].closest('a') || arguments[0];", icon)
            if anchor:
                self.driver.execute_script("arguments[0].click();", anchor)
                print("  [OK] 테이블 다운로드 아이콘 클릭")
                return True
            print("  [WARNING] 다운로드 아이콘의 앵커를 찾지 못함")
            return False
        except Exception as e:
            print(f"  [INFO] 테이블 다운로드 아이콘을 찾지 못함: {e}")
            return False

    def _go_back_to_list(self):
        """상세 화면에서 검색 결과 리스트로 돌아간다 (검색 결과로 돌아가기 버튼)"""
        try:
            button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button .fds-icon-name-arrow-back"))
            )
            # 버튼 객체 찾기
            btn = self.driver.execute_script("""
            const icons = document.querySelectorAll('.fds-icon-name-arrow-back');
            for (let icon of icons) {
                let btn = icon.closest('button');
                if (btn) return btn;
            }
            return null;
            """)
            if btn:
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(3)
                
                # 리스트 페이지 로드 완료 대기 (JavaScript 데이터 로드)
                for i in range(45):
                    page_source = self.driver.page_source
                    if 'statementsAndInvoicesPayload' in page_source:
                        time.sleep(2)
                        print("  [OK] 검색 결과로 돌아감 (페이지 로드 완료)")
                        return True
                    time.sleep(1)
                
                print("  [OK] 검색 결과로 돌아감 (로드 신호 미수신)")
                return True
            return False
        except Exception as e:
            print(f"  [INFO] 돌아가기 버튼 클릭 실패: {e}")
            return False
    
    def download_statement(self, payment_info):
        """
        특정 명세서 다운로드
        
        Args:
            payment_info: dict - 명세서 정보 {'paymentRequestId', 'invoiceId', 'amountProcessed', 'datePaid' 등}
            
        Returns:
            bool: 다운로드 성공 여부
        """
        payment_request_id = payment_info.get('paymentRequestId')
        invoice_id = payment_info.get('invoiceId')
        amount = int(payment_info.get('amountProcessed', 0))
        date_paid = payment_info.get('datePaid', '').replace('-', '')  # 20251229 형식
        file_path = payment_info.get('paymentRequestFilePath', '')
        notice_path = payment_info.get('paymentNoticePath', '')
        
        # datePaid가 없으면 skip
        if not date_paid:
            print(f"  [SKIP] 결제 미완료: {payment_request_id}")
            return False
        
        print(f"  명세서: {payment_request_id}, 금액: {amount}, 결제일: {date_paid}")
        
        try:
            # 현재 detail view에 있으면 list로 돌아가기
            current_url = self.driver.current_url
            if 'invoiceDetails' in current_url or 'remittanceDetails' in current_url:
                self._go_back_to_list()

            before_url = self.driver.current_url

            # 인보이스 버튼 클릭: 현재 페이지→다음 페이지 순회 (최대 15페이지)
            found = False
            for page_idx in range(15):
                if self._click_invoice_button_on_current_page(invoice_id):
                    print("  [OK] 결제 ID 버튼 클릭")
                    found = True
                    break
                if not self._go_to_next_page():
                    break
            if not found:
                print(f"  [ERROR] 지불 ID 버튼을 찾지 못함: invoiceId={invoice_id}")
                return False

            # 상세 화면 로드까지 대기 후 진행
            if not self._wait_for_detail_page(before_url):
                print("  [WARNING] 상세 화면 로드 신호를 받지 못함 (계속 시도)")

            windows = self.driver.window_handles
            if len(windows) > 1:
                self.driver.switch_to.window(windows[-1])
                print("  [OK] 상세 탭으로 전환")
            else:
                print("  [INFO] 동일 탭에서 상세 표시")

            # 페이지 완전 로드 대기
            time.sleep(3)

            # 상세 화면에서 바로 다운로드 아이콘 클릭
            if not self._click_table_download_icon():
                print("  [ERROR] 다운로드 아이콘을 클릭하지 못함")
                self._go_back_to_list()
                return False

            # 아이콘 클릭 후 다운로드 대기
            time.sleep(4)
            
            # 다운로드된 파일 확인 및 이동
            temp_dir = Path(self.temp_download_dir)
            # 최신 다운로드 파일 하나 선택 (우선 txt/csv)
            candidates = list(temp_dir.glob('*.txt')) + list(temp_dir.glob('*.csv'))
            if not candidates:
                candidates = [p for p in temp_dir.iterdir() if p.is_file()]
            files = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
            
            if files:
                source_file = files[0]
                # 원본 확장자 확인 후 CSV로 변경
                source_ext = source_file.suffix if source_file.suffix else ''
                target_ext = '.csv' if source_ext.lower() in {'.txt', ''} else source_ext
                final_filename = f"익스피디아_{date_paid}_{amount}{target_ext}"
                final_path = self.download_dir / final_filename
                if final_path.exists():
                    final_path.unlink()
                
                source_file.rename(final_path)
                print(f"  [OK] 파일 저장: {final_filename}")
                
                # 다운로드 완료 후 리스트 페이지로 돌아가기
                self._go_back_to_list()
                
                return True
            else:
                print(f"  [WARNING] 다운로드 파일을 찾을 수 없음")
                # 디버깅: temp_downloads의 모든 파일 확인
                all_files = [p.name for p in temp_dir.iterdir() if p.is_file()]
                if all_files:
                    print(f"  [DEBUG] temp_downloads 파일: {all_files[:10]}")
                time.sleep(2)
                
                # 재시도: 추가 대기 후 다시 확인
                candidates = list(temp_dir.glob('*.txt')) + list(temp_dir.glob('*.csv'))
                if not candidates:
                    candidates = [p for p in temp_dir.iterdir() if p.is_file()]
                files = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
                
                if files:
                    source_file = files[0]
                    source_ext = source_file.suffix if source_file.suffix else ''
                    target_ext = '.csv' if source_ext.lower() in {'.txt', ''} else source_ext
                    final_filename = f"익스피디아_{date_paid}_{amount}{target_ext}"
                    final_path = self.download_dir / final_filename
                    if final_path.exists():
                        final_path.unlink()
                    
                    source_file.rename(final_path)
                    print(f"  [OK] 파일 저장 (재시도): {final_filename}")
                    self._go_back_to_list()
                    return True
                else:
                    print(f"  [WARNING] 재시도에도 파일을 찾지 못함 (리스트로 돌아감)")
                    self._go_back_to_list()
                    return False
                
        except Exception as e:
            print(f"  [ERROR] 다운로드 오류: {e}")
            import traceback
            traceback.print_exc()
            # 예외 발생 시에도 리스트로 돌아가기
            try:
                self._go_back_to_list()
            except:
                pass
            return False
            
        except Exception as e:
            print(f"    [ERROR] 다운로드 중 오류: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def download_statements(self, limit=None, payment_ids=None, date_from=None, date_to=None):
        """
        여러 명세서 다운로드
        
        Args:
            limit: 다운로드할 최대 개수 (None이면 제한 없음)
            payment_ids: 특정 ID 리스트 (None이면 전체)
            date_from: 시작 날짜 (YYYY-MM-DD 형식, None이면 1년 전)
            date_to: 종료 날짜 (YYYY-MM-DD 형식, None이면 오늘)
            
        Returns:
            int: 다운로드 성공한 파일 개수
        """
        from datetime import datetime, timedelta
        
        # 기본 날짜 설정 (결제날짜 기준)
        if date_to is None:
            date_to = datetime.now().strftime('%Y-%m-%d')
        if date_from is None:
            date_from = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        print(f"\n[다운로드] 명세서 일괄 다운로드 시작...")
        print(f"  기간: {date_from} ~ {date_to} (결제날짜 기준)")
        
        statements = self.get_statement_list()
        
        if not statements:
            print("  [ERROR] 다운로드할 명세서가 없습니다")
            return 0
        
        # 날짜 필터링 (datePaid 기준)
        filtered_statements = []
        for stmt in statements:
            date_paid = stmt.get('datePaid', '')
            if date_paid and date_from <= date_paid <= date_to:
                filtered_statements.append(stmt)
        
        print(f"  기간 내 명세서: {len(filtered_statements)}개")
        
        # 추가 필터링
        if payment_ids:
            filtered_statements = [s for s in filtered_statements if s.get('paymentRequestId') in payment_ids]
        
        # 엑셀에 있지만 파일이 없는 명세서 필터링
        excel_path = self.base_dir / '매출 및 입금 결과.xlsx'
        excel_records = set()
        missing_from_excel = set()
        
        if excel_path.exists():
            try:
                from openpyxl import load_workbook
                wb_check = load_workbook(excel_path, read_only=True)
                if '익스피디아' in wb_check.sheetnames:
                    ws_check = wb_check['익스피디아']
                    for row in ws_check.iter_rows(min_row=2, values_only=True):
                        if row[2] and row[3]:  # 결제날짜, 처리금액
                            date_str = str(row[2]).replace('-', '')
                            amount_str = str(row[3]).replace(',', '').replace('.0', '').strip()
                            excel_records.add(f"{date_str}_{amount_str}")
                    wb_check.close()
                    print(f"  엑셀 명세서 기록: {len(excel_records)}개")
            except Exception as e:
                print(f"  [WARN] 엑셀 읽기 실패: {e}")
        
        # 이미 저장된 파일 확인 (중복 다운로드 방지)
        existing_files = set()
        for f in self.download_dir.glob('익스피디아_*'):
            # 익스피디아_20251229_4793717.csv 형식에서 date_amount 추출
            parts = f.stem.split('_')  # 익스피디아, 20251229, 4793717
            if len(parts) >= 3:
                date_amount = f"{parts[1]}_{parts[2]}"
                existing_files.add(date_amount)
        
        # 엑셀에는 있지만 파일이 없는 항목 계산
        missing_from_excel = excel_records - existing_files
        if missing_from_excel:
            print(f"  엑셀에 있지만 파일이 없는 명세서: {len(missing_from_excel)}개")
        
        # 이미 저장된 파일은 필터링 (엑셀 우선순위)
        statements_to_download = []
        for stmt in filtered_statements:
            amount = int(stmt.get('amountProcessed', 0))
            date_paid = stmt.get('datePaid', '').replace('-', '')
            date_amount = f"{date_paid}_{amount}"
            
            # 우선순위: 엑셀에는 있지만 파일이 없는 것
            if date_amount in missing_from_excel:
                statements_to_download.append(stmt)
            # 파일이 아예 없는 경우도 추가
            elif date_amount not in existing_files:
                statements_to_download.append(stmt)
            else:
                payment_id = stmt.get('paymentRequestId', '')
                print(f"  [SKIP] 이미 저장됨: {payment_id} (날짜: {date_paid}, 금액: {amount})")
        
        if limit:
            statements_to_download = statements_to_download[:limit]
        
        print(f"  다운로드 대상: {len(statements_to_download)}개 명세서")
        
        downloaded_count = 0
        for i, stmt in enumerate(statements_to_download, 1):
            print(f"\n[{i}/{len(statements_to_download)}] 다운로드 중...")
            
            success = self.download_statement(stmt)
            if success:
                downloaded_count += 1
            
            # 딜레이
            if i < len(statements_to_download):
                time.sleep(2)
        
        print(f"\n[완료] 총 {downloaded_count}/{len(statements_to_download)}개 파일 다운로드 성공")
        
        # 다운로드 후 엑셀 업데이트
        self._update_excel_with_statements(filtered_statements)
        
        return downloaded_count
    
    def _update_excel_with_statements(self, statements):
        """
        다운로드한 명세서 정보를 '매출 및 입금결과.xlsx'의 '익스피디아' 시트에 추가
        
        Args:
            statements: 명세서 정보 리스트 (다운로드 대상 전체)
        """
        try:
            from openpyxl import load_workbook, Workbook
            from datetime import datetime
            
            excel_path = self.base_dir / '매출 및 입금 결과.xlsx'
            
            print(f"\n[엑셀 경로 확인]")
            print(f"  base_dir: {self.base_dir}")
            print(f"  excel_path: {excel_path}")
            
            # 폴더 내 엑셀 파일 목록 확인
            excel_files = list(self.base_dir.glob('*.xlsx'))
            print(f"  폴더 내 엑셀 파일 ({len(excel_files)}개):")
            for f in excel_files[:10]:
                print(f"    - {f.name}")
            
            print(f"  파일 존재 여부: {excel_path.exists()}")
            
            # 파일이 없으면 새로 생성
            if not excel_path.exists():
                print(f"\n[엑셀 생성] {excel_path}")
                wb = Workbook()
                # 기본 시트 이름 변경
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
                ws = wb.create_sheet('익스피디아')
                # 헤더 추가
                ws.append(['요청날짜', '지불ID', '결제날짜', '처리금액'])
                wb.save(excel_path)
                print(f"  [OK] 새 엑셀 파일 생성 완료")
            else:
                print(f"\n[엑셀 업데이트] 기존 파일 사용")
            
            # 엑셀 파일 로드
            wb = load_workbook(excel_path)
            
            # 익스피디아 시트 확인 또는 생성
            if '익스피디아' not in wb.sheetnames:
                ws = wb.create_sheet('익스피디아')
                # 헤더 추가
                ws.append(['요청날짜', '지불ID', '결제날짜', '처리금액'])
                print("  [INFO] '익스피디아' 시트 생성 및 헤더 추가")
            else:
                ws = wb['익스피디아']
            
            # 기존 지불ID 목록 수집 (B열)
            existing_payment_ids = set()
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1]:  # B열 (지불ID)
                    existing_payment_ids.add(str(row[1]).strip())
            
            print(f"  기존 데이터: {len(existing_payment_ids)}건")
            
            # 새 데이터 추가
            added_count = 0
            for stmt in statements:
                payment_id = str(stmt.get('paymentRequestId', '')).strip()
                
                if not payment_id or payment_id in existing_payment_ids:
                    continue
                
                # 데이터 추출
                date_requested = stmt.get('dateRequested', '')
                date_paid = stmt.get('datePaid', '')
                amount = stmt.get('amountProcessed', 0)
                
                # 금액을 정수로 변환 후 쉼표 추가 (천 단위 구분)
                amount_int = int(amount) if isinstance(amount, (int, float)) else 0
                amount_formatted = f"{amount_int:,}"
                
                # 행 추가
                ws.append([date_requested, payment_id, date_paid, amount_formatted])
                existing_payment_ids.add(payment_id)
                added_count += 1
            
            print(f"  새로 추가: {added_count}건")
            
            # 요청날짜로 정렬 (최근 날짜 먼저, 헤더 제외)
            if ws.max_row > 1:
                data_rows = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    data_rows.append(list(row))
                
                # 요청날짜(A열, 인덱스 0) 기준 내림차순 정렬
                data_rows.sort(key=lambda x: x[0] if x[0] else '', reverse=True)
                
                # 기존 데이터 삭제 (헤더 제외)
                ws.delete_rows(2, ws.max_row - 1)
                
                # 정렬된 데이터 다시 추가
                for row_data in data_rows:
                    ws.append(row_data)
                
                print(f"  정렬 완료: 요청날짜 기준 최근순")
            
            # 저장
            wb.save(excel_path)
            wb.close()
            print(f"  [OK] 엑셀 저장 완료")
            
        except Exception as e:
            print(f"  [ERROR] 엑셀 업데이트 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def close(self):
        """브라우저 종료"""
        if self.driver:
            self.driver.quit()
            print("\n브라우저 종료")


if __name__ == '__main__':
    # 간단한 사용 예시
    downloader = ExpediaDownloader()
    downloader.setup_driver()
    
    if downloader.login():
        downloader.navigate_to_statements()
        statements = downloader.get_statement_list()
        
        if statements:
            print(f"\n총 {len(statements)}개 명세서 발견")
            # 전체 다운로드 (엑셀 기반 필터링으로 필요한 것만 다운로드)
            downloader.download_statements()
    
    downloader.close()
