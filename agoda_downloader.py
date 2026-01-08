"""
아고다(Agoda) 명세서 다운로더
- Partner Central에서 명세서 목록 조회 및 CSV 다운로드
- Excel 파일에 자동으로 데이터 업데이트 (중복 제거, 자동 정렬)
"""

import os
import sys
import json
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional
from dataclasses import dataclass
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# .env 파일 로드
load_dotenv()


# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class RemittanceRecord:
    """아고다 명세서 레코드"""
    remittance_info_id: str  # 고유 ID
    date: str  # 요청 날짜 (예: 07-Jan-2026)
    currency: str  # 통화 (KRW)
    amount: float  # 금액
    payout_id: str  # 지불 ID (예: 20251229-49798)
    payout_method: str  # 지불 방법
    

class AgodaDownloader:
    """아고다 명세서 다운로더"""
    
    def __init__(self, base_dir: str = None, download_dir: str = None, cookies_file: str = None):
        """
        Args:
            base_dir: 기본 디렉토리 (기본값: 스크립트 디렉토리)
            download_dir: 다운로드 디렉토리 (기본값: base_dir/ota-adjustment)
            cookies_file: 쿠키 저장 파일 (기본값: base_dir/agoda_cookies.json)
        """
        # 기본 경로 설정
        if base_dir is None:
            base_dir = str(Path(__file__).parent)
        self.base_dir = Path(base_dir)
        
        # 다운로드 디렉토리 설정
        if download_dir is None:
            download_dir = self.base_dir / 'ota-adjustment'
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)

        # 크롬 프로필 디렉토리 (세션 유지용)
        self.profile_dir = self.base_dir / 'chrome-profile'
        self.profile_dir.mkdir(parents=True, exist_ok=True)
        
        # 쿠키 파일 경로 설정
        if cookies_file is None:
            cookies_file = self.base_dir / 'agoda_cookies.json'
        self.cookies_file = Path(cookies_file)
        
        # 환경 변수에서 credentials 가져오기
        self.username = os.environ.get('AGODA_USERNAME')
        self.password = os.environ.get('AGODA_PASSWORD')
        
        # 명세서 페이지 URL (환경변수로 커스터마이징 가능)
        self.remittances_url = os.environ.get(
            'AGODA_REMITTANCES_URL',
            'https://ycs.agoda.com/en-us/Finance/Remittances/1709863'
        )
        
        self.driver = None
        self.wait = None
        
        logger.info(f"AgodaDownloader 초기화")
        logger.info(f"  Base dir: {self.base_dir}")
        logger.info(f"  Download dir: {self.download_dir}")
        logger.info(f"  Cookies file: {self.cookies_file}")
        logger.info(f"  Chrome profile dir: {self.profile_dir}")
        logger.info(f"  Remittances URL: {self.remittances_url}")
    
    def setup_driver(self):
        """Chrome WebDriver 설정"""
        logger.info("Chrome WebDriver 설정 중...")
        
        options = webdriver.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-blink-features=AutomationControlled')
        
        # 다운로드 설정
        prefs = {
            'download.default_directory': str(self.download_dir),
            'download.prompt_for_download': False,
            'profile.default_content_settings.popups': 0
        }
        options.add_experimental_option('prefs', prefs)
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        self.wait = WebDriverWait(self.driver, 10)
        logger.info("WebDriver 설정 완료")
    
    def load_cookies(self) -> bool:
        """저장된 쿠키 로드"""
        if self.cookies_file.exists():
            try:
                with open(self.cookies_file, 'r') as f:
                    cookies = json.load(f)
                
                # 쿠키 추가
                for cookie in cookies:
                    # 필요한 필드만 유지
                    cookie_dict = {
                        'name': cookie.get('name'),
                        'value': cookie.get('value'),
                        'domain': cookie.get('domain', '.agoda.com'),
                        'path': cookie.get('path', '/'),
                    }
                    try:
                        self.driver.add_cookie(cookie_dict)
                    except Exception as e:
                        logger.warning(f"쿠키 추가 실패: {cookie.get('name')} - {e}")
                
                logger.info(f"쿠키 로드 완료: {len(cookies)}개")
                return True
            except Exception as e:
                logger.warning(f"쿠키 로드 실패: {e}")
                return False
        return False
    
    def save_cookies(self):
        """쿠키 저장"""
        try:
            cookies = self.driver.get_cookies()
            with open(self.cookies_file, 'w') as f:
                json.dump(cookies, f, indent=2)
            logger.info(f"쿠키 저장 완료: {len(cookies)}개")
        except Exception as e:
            logger.error(f"쿠키 저장 실패: {e}")
    
    def login(self):
        """아고다 로그인 (iframe 처리)"""
        try:
            logger.info("아고다 로그인 시작...")
            
            # 로그인 페이지 방문
            self.driver.get('http://ycs.agoda.com/mldc/en-us/public/login')
            time.sleep(3)
            
            # 쿠키 먼저 시도
            if self.load_cookies():
                self.driver.get(self.remittances_url)
                time.sleep(2)
                
                # 로그인 상태 확인 (테이블 요소로 확인)
                try:
                    self.wait.until(EC.presence_of_element_located((By.ID, 'tblRemittances')))
                    logger.info("저장된 쿠키로 로그인 성공")
                    return True
                except:
                    logger.info("저장된 쿠키가 만료됨, 새로 로그인...")
            
            # 아이디/비밀번호로 로그인
            if not self.username or not self.password:
                raise ValueError("AGODA_USERNAME, AGODA_PASSWORD 환경 변수 필요")
            
            # 페이지가 동적으로 로드되므로 더 길게 대기
            logger.info("로그인 페이지 로드 대기 중...")
            time.sleep(3)
            
            # iframe이 있는지 확인 (로그인 폼이 iframe 내부에 있음)
            wait_long = WebDriverWait(self.driver, 20)
            
            # iframe 대기
            iframe = wait_long.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'iframe[data-cy="ul-app-frame"]'))
            )
            logger.info("iframe 발견, iframe 내부로 전환 중...")
            self.driver.switch_to.frame(iframe)
            
            # iframe 내에서 이메일 입력 필드 찾기
            email_input = wait_long.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-cy="unified-email-input"]'))
            )
            email_input.clear()
            email_input.send_keys(self.username)
            logger.info("이메일 입력 완료")
            
            # "진행하기" 버튼 클릭
            continue_button = wait_long.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-cy="unified-email-continue-button"]'))
            )
            continue_button.click()
            logger.info("이메일 확인 버튼 클릭")
            
            # 비밀번호 또는 2차 인증 단계 대기
            logger.info("비밀번호/2차 인증 단계 대기 중... 필요 시 브라우저에서 직접 인증을 완료하세요.")
            print("브라우저에서 비밀번호/2차 인증을 완료한 후 Enter를 눌러주세요 (창을 닫지 마세요).")
            input()

            # iframe에서 벗어나기
            self.driver.switch_to.default_content()
            logger.info("iframe에서 벗어남")
            
            # 로그인 완료 확인 및 명세서 페이지 접근 재시도
            for attempt in range(3):
                logger.info(f"명세서 페이지 로드 시도 {attempt + 1}/3 ...")
                self.driver.get(self.remittances_url)
                try:
                    WebDriverWait(self.driver, 20).until(
                        EC.presence_of_element_located((By.ID, 'tblRemittances'))
                    )
                    logger.info("로그인 성공")
                    self.save_cookies()
                    return True
                except TimeoutException:
                    logger.warning("명세서 테이블이 보이지 않음. 브라우저에서 로그인 완료 여부를 확인 후 Enter.")
                    print("브라우저에서 로그인 단계가 모두 끝났는지 확인 후 Enter를 눌러 재시도하세요.")
                    input()
            
            raise TimeoutException("로그인 후 명세서 페이지 로드 실패")
            
        except Exception as e:
            logger.error(f"로그인 실패: {e}")
            raise
    
    def get_remittance_list(self) -> List[RemittanceRecord]:
        """명세서 목록 조회"""
        try:
            logger.info("명세서 목록 조회 중...")
            
            remittances = []
            table = self.wait.until(
                EC.presence_of_element_located((By.ID, 'tblRemittances'))
            )
            
            # 테이블의 모든 행 처리
            rows = table.find_elements(By.XPATH, './/tbody/tr[contains(@id, "cardInfo") = false]')
            logger.info(f"발견된 명세서 행: {len(rows)}개")
            
            for row in rows:
                try:
                    # 행 ID가 remittanceInfoId
                    row_id = row.get_attribute('id')
                    if not row_id or 'cardInfo' in row_id or 'trAdditional' in row_id:
                        continue
                    
                    # 각 셀에서 데이터 추출
                    cells = row.find_elements(By.TAG_NAME, 'td')
                    
                    if len(cells) < 9:
                        continue
                    
                    # 컬럼 순서: checkbox, date, currency, amount, payoutId, fapiao1, fapiao2, payoutMethod, statement
                    date_str = cells[1].text.strip()  # "07-Jan-2026"
                    currency = cells[2].text.strip()  # "KRW"
                    amount_str = cells[3].text.strip()  # "15,123,015.00"
                    payout_id = cells[4].text.strip()  # "20251229-49798"
                    payout_method = cells[7].text.strip()  # "Agoda Card" or "Telex Transfer"
                    
                    # 금액 파싱 (쉼표 제거)
                    try:
                        amount = float(amount_str.replace(',', ''))
                    except ValueError:
                        logger.warning(f"금액 파싱 실패: {amount_str}")
                        continue
                    
                    record = RemittanceRecord(
                        remittance_info_id=row_id,
                        date=date_str,
                        currency=currency,
                        amount=amount,
                        payout_id=payout_id,
                        payout_method=payout_method
                    )
                    remittances.append(record)
                    
                except Exception as e:
                    logger.warning(f"행 처리 실패: {e}")
                    continue
            
            logger.info(f"명세서 조회 완료: {len(remittances)}개")
            return remittances
            
        except Exception as e:
            logger.error(f"명세서 목록 조회 실패: {e}")
            raise
    
    def download_remittance(self, remittance: RemittanceRecord) -> bool:
        """개별 명세서 다운로드 (Export to Excel)
        
        Returns:
            다운로드 성공 여부
        """
        try:
            logger.info(f"명세서 다운로드: {remittance.payout_id} ({remittance.amount:,.0f} {remittance.currency})")
            
            # 파일명 생성: 아고다_결제금액_요청날짜.csv
            # 날짜 형식 변환: "07-Jan-2026" → "20260107"
            try:
                date_obj = datetime.strptime(remittance.date, '%d-%b-%Y')
                date_str = date_obj.strftime('%Y%m%d')
            except ValueError:
                date_str = remittance.date.replace('-', '')
            
            amount_int = int(remittance.amount)
            target_filename = f"아고다_{date_str}_{amount_int}.csv"
            target_file = self.download_dir / target_filename
            
            # 이미 존재하면 스킵
            if target_file.exists():
                logger.info(f"파일 이미 존재: {target_filename}")
                return False
            
            # 행을 클릭해서 상세 페이지 열기
            row = self.driver.find_element(By.ID, remittance.remittance_info_id)
            row.click()
            time.sleep(2)
            
            # 하단으로 스크롤해서 Export to Excel 버튼 찾기
            export_button = self.wait.until(
                EC.presence_of_element_located((By.ID, 'btnExport'))
            )
            
            # 버튼까지 스크롤
            self.driver.execute_script("arguments[0].scrollIntoView(true);", export_button)
            time.sleep(1)
            
            # 현재 파일 개수 기록 (다운로드된 새 파일 감지용)
            existing_files_before = set(self.download_dir.glob('Remittances*.xlsx'))
            
            # 버튼 클릭
            export_button.click()
            
            # 파일 다운로드 대기
            time.sleep(3)
            max_wait = 10
            downloaded_file = None
            
            while max_wait > 0:
                # 새로 생성된 파일이 있는지 확인
                current_files = set(self.download_dir.glob('Remittances*.xlsx'))
                new_files = current_files - existing_files_before
                if new_files:
                    downloaded_file = max(new_files, key=lambda p: p.stat().st_mtime)
                    logger.info(f"다운로드된 파일: {downloaded_file.name}")
                    break
                
                time.sleep(1)
                max_wait -= 1
            
            if not downloaded_file:
                logger.warning(f"파일 다운로드 타임아웃: {target_filename}")
                return False
            
            # xlsx 파일을 csv로 변환하고 이름 변경
            try:
                import pandas as pd
                df = pd.read_excel(downloaded_file)
                df.to_csv(target_file, index=False, encoding='utf-8-sig')
                logger.info(f"CSV 변환 완료: {target_filename}")
                
                # 원본 xlsx 파일 삭제
                downloaded_file.unlink()
                return True
            except Exception as e:
                logger.error(f"CSV 변환 실패: {e}")
                # 변환 실패해도 xlsx 파일을 원하는 이름으로 변경
                try:
                    xlsx_target = self.download_dir / target_filename.replace('.csv', '.xlsx')
                    downloaded_file.rename(xlsx_target)
                    logger.info(f"파일 이름 변경: {xlsx_target.name}")
                    return True
                except Exception as rename_err:
                    logger.error(f"파일 이름 변경 실패: {rename_err}")
                    return False
            
        except Exception as e:
            logger.error(f"명세서 다운로드 실패: {remittance.payout_id} - {e}")
            return False
    
    def download_remittances(self, start_date: str = None, end_date: str = None) -> List[RemittanceRecord]:
        """명세서 다운로드 (날짜 범위)
        
        기본값: 지난 1년치 데이터만 다운로드
        
        Args:
            start_date: 시작 날짜 (YYYY-MM-DD 형식, 기본값: 오늘 - 1년)
            end_date: 종료 날짜 (YYYY-MM-DD 형식, 기본값: 오늘)
        
        Returns:
            조회된 명세서 목록 (다운로드 성공 여부와 관계없이)
        """
        try:
            # 기본값: 과거 1년부터 오늘까지
            if not start_date:
                start_date_obj = datetime.now().replace(year=datetime.now().year - 1)
                start_date = start_date_obj.strftime('%Y-%m-%d')
            if not end_date:
                end_date = datetime.now().strftime('%Y-%m-%d')
            
            logger.info(f"조회 범위: {start_date} ~ {end_date}")
            
            remittances = self.get_remittance_list()
            
            # 디버깅: 조회된 명세서의 날짜 범위 확인
            if remittances:
                dates = []
                for r in remittances:
                    try:
                        r_date = datetime.strptime(r.date, '%d-%b-%Y')
                        dates.append(r_date)
                    except ValueError:
                        pass
                if dates:
                    min_date = min(dates).strftime('%Y-%m-%d')
                    max_date = max(dates).strftime('%Y-%m-%d')
                    logger.info(f"조회된 명세서 날짜 범위: {min_date} ~ {max_date}")
            
            # 날짜 필터링
            filtered = []
            for r in remittances:
                # 아고다 날짜 형식: "07-Jan-2026"
                try:
                    r_date = datetime.strptime(r.date, '%d-%b-%Y').strftime('%Y-%m-%d')
                    
                    if start_date and r_date < start_date:
                        continue
                    if end_date and r_date > end_date:
                        continue
                    filtered.append(r)
                except ValueError:
                    logger.warning(f"날짜 파싱 실패: {r.date}")
            
            remittances = filtered
            logger.info(f"필터링된 명세서: {len(remittances)}개 (필터: {start_date} ~ {end_date})")
            
            # 엑셀에 있지만 파일이 없는 명세서만 다운로드
            excel_path = self.base_dir / '매출 및 입금 결과.xlsx'
            if excel_path.exists():
                try:
                    wb_check = load_workbook(excel_path, read_only=True)
                    if '아고다' in wb_check.sheetnames:
                        ws_check = wb_check['아고다']
                        excel_records = set()
                        for row in ws_check.iter_rows(min_row=2, values_only=True):
                            if row[0] and row[1]:  # 날짜, 금액
                                date_str = str(row[0]).replace('-', '')
                                amount_str = str(row[1]).replace(',', '').replace('.0', '').strip()
                                excel_records.add(f"{date_str}_{amount_str}")
                        wb_check.close()
                        
                        # 실제 파일 확인
                        existing_files = set()
                        for f in self.download_dir.glob('아고다_*.csv'):
                            parts = f.stem.split('_')
                            if len(parts) >= 3:
                                existing_files.add(f"{parts[1]}_{parts[2]}")
                        
                        # 엑셀에는 있지만 파일이 없는 항목 필터링
                        missing_records = excel_records - existing_files
                        if missing_records:
                            logger.info(f"엑셀에 있지만 파일이 없는 명세서: {len(missing_records)}개")
                            remittances_to_download = []
                            for r in remittances:
                                date_obj = datetime.strptime(r.date, '%d-%b-%Y')
                                date_str = date_obj.strftime('%Y%m%d')
                                amount_int = int(r.amount)
                                key = f"{date_str}_{amount_int}"
                                if key in missing_records:
                                    remittances_to_download.append(r)
                            remittances = remittances_to_download
                            logger.info(f"다운로드 대상(엑셀 기반): {len(remittances)}개")
                except Exception as e:
                    logger.warning(f"엑셀 필터링 실패, 전체 다운로드: {e}")
            
            # 개별 다운로드
            for remittance in remittances:
                self.download_remittance(remittance)
            
            # 조회된 명세서 모두 반환 (Excel 업데이트용)
            return remittances
            
        except Exception as e:
            logger.error(f"명세서 다운로드 실패: {e}")
            raise
    
    def _update_excel_with_remittances(self, remittances: List[RemittanceRecord]):
        """Excel 파일에 명세서 업데이트 (CSV 다운로드 여부와 관계없이)"""
        try:
            excel_path = self.base_dir / '매출 및 입금 결과.xlsx'
            logger.info(f"Excel 파일 업데이트: {excel_path}")
            logger.info(f"업데이트할 명세서: {len(remittances)}개")
            
            # Excel 파일 로드 또는 생성
            if excel_path.exists():
                wb = load_workbook(excel_path)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                # 기본 시트 제거
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # "아고다" 시트 생성 또는 선택
            sheet_name = '아고다'
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # 헤더 확인 및 설정 (항상 1행에)
            if ws.cell(row=1, column=1).value != '요청날짜':
                ws.cell(row=1, column=1, value='요청날짜')
                ws.cell(row=1, column=2, value='처리금액')
                ws.cell(row=1, column=3, value='지불ID')
            
            # 기존 데이터 읽기 (중복 제거용) - 2행부터 시작
            existing_payout_ids = set()
            for row_idx in range(2, ws.max_row + 1):
                payout_id = ws.cell(row=row_idx, column=3).value
                if payout_id:
                    existing_payout_ids.add(str(payout_id))
            
            logger.info(f"기존 지불ID: {len(existing_payout_ids)}개")
            
            # 새로운 명세서만 추가
            added_count = 0
            for remittance in remittances:
                if remittance.payout_id not in existing_payout_ids:
                    # 날짜를 YYYY-MM-DD 형식으로 변환
                    try:
                        date_obj = datetime.strptime(remittance.date, '%d-%b-%Y')
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        date_str = remittance.date
                    
                    # 금액을 정수로 표시 (소수점 제거, 쉼표 구분자 추가)
                    amount_int = int(remittance.amount)
                    amount_formatted = f"{amount_int:,}"
                    
                    # 새 행 추가 (2행부터 시작)
                    new_row = ws.max_row + 1
                    if new_row == 1:  # 비어있으면 헤더 다음
                        new_row = 2
                    
                    ws.cell(row=new_row, column=1, value=date_str)
                    ws.cell(row=new_row, column=2, value=amount_formatted)
                    ws.cell(row=new_row, column=3, value=remittance.payout_id)
                    
                    added_count += 1
                    logger.info(f"  추가: {remittance.payout_id} - {amount_int:,} KRW ({date_str})")
            
            logger.info(f"추가된 명세서: {added_count}개")
            
            # 요청날짜로 정렬 (최근 날짜 먼저, 헤더 제외)
            if ws.max_row > 1:
                data_rows = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    data_rows.append(list(row))
                
                # 요청날짜(A열, 인덱스 0) 기준 내림차순 정렬
                data_rows.sort(key=lambda x: x[0] if x[0] else '', reverse=True)
                
                # 기존 데이터 삭제 (헤더 제외)
                ws.delete_rows(2, ws.max_row - 1)
                
                # 정렬된 데이터 다시 추가
                for row_data in data_rows:
                    ws.append(row_data)
                
                logger.info(f"정렬 완료: 요청날짜 기준 최근순")
            
            # 파일 저장
            wb.save(excel_path)
            logger.info(f"Excel 파일 업데이트 완료: {excel_path}")
            
        except Exception as e:
            logger.error(f"Excel 파일 업데이트 실패: {e}")
            raise
    
    def close(self):
        """WebDriver 종료"""
        if self.driver:
            self.driver.quit()
            logger.info("WebDriver 종료")
    
    def run(self, start_date: str = None, end_date: str = None):
        """전체 실행 (로그인 → 목록 조회 → 다운로드 → Excel 업데이트)
        
        Args:
            start_date: 시작 날짜 (YYYY-MM-DD 형식, optional)
            end_date: 종료 날짜 (YYYY-MM-DD 형식, optional)
        """
        try:
            self.setup_driver()
            self.login()
            
            # 명세서 다운로드
            remittances = self.download_remittances(start_date, end_date)
            
            # Excel 업데이트
            if remittances:
                self._update_excel_with_remittances(remittances)
            else:
                logger.info("업데이트할 명세서 없음")
            
            logger.info("아고다 명세서 다운로드 완료")
            
        except Exception as e:
            logger.error(f"실행 중 오류: {e}")
            raise
        finally:
            # 환경변수 AGODA_KEEP_BROWSER_OPEN=1 이면 브라우저를 닫지 않음
            keep_open = os.environ.get('AGODA_KEEP_BROWSER_OPEN', '0')
            if keep_open.lower() not in ('1', 'true', 'yes'):
                self.close()
            else:
                logger.info("AGODA_KEEP_BROWSER_OPEN 설정으로 브라우저를 열어둡니다.")


if __name__ == '__main__':
    """명령행 실행"""
    import argparse
    
    parser = argparse.ArgumentParser(description='아고다 명세서 다운로더')
    parser.add_argument('--start-date', help='시작 날짜 (YYYY-MM-DD)')
    parser.add_argument('--end-date', help='종료 날짜 (YYYY-MM-DD)')
    parser.add_argument('--base-dir', help='기본 디렉토리')
    
    args = parser.parse_args()
    
    downloader = AgodaDownloader(base_dir=args.base_dir)
    downloader.run(start_date=args.start_date, end_date=args.end_date)
