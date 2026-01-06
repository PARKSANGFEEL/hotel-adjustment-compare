# -*- coding: utf-8 -*-
"""
부킹닷컴(Booking.com) 명세서 자동 다운로더

기능:
1. 부킹 Partner Central 로그인 (쿠키 기반)
2. 재무 → 대금지급정보 페이지 접근
3. 명세서 테이블 파싱
4. CSV 다운로드 (자동)
5. 매출 및 입금 결과.xlsx에 데이터 추가

작성자: GitHub Copilot
날짜: 2026-01-07
"""

import os
import json
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Dict
import pandas as pd
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import re


# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class BookingDownloader:
    """부킹닷컴 명세서 자동 다운로더"""
    
    def __init__(self, base_dir: str = None, download_dir: str = None, cookies_file: str = None):
        """
        초기화
        
        Args:
            base_dir: 기본 디렉토리 (기본값: 스크립트 디렉토리)
            download_dir: 다운로드 디렉토리 (기본값: base_dir/ota-adjustment)
            cookies_file: 쿠키 저장 파일 (기본값: base_dir/booking_cookies.json)
        """
        if base_dir is None:
            base_dir = str(Path(__file__).parent)
        self.base_dir = Path(base_dir)
        
        if download_dir is None:
            download_dir = self.base_dir / 'ota-adjustment'
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        
        if cookies_file is None:
            cookies_file = self.base_dir / 'booking_cookies.json'
        self.cookies_file = Path(cookies_file)
        
        # 환경변수에서 credentials 가져오기
        self.username = os.environ.get('BOOKING_USERNAME')
        self.password = os.environ.get('BOOKING_PASSWORD')
        
        # 부킹 호텔 ID (환경변수에서 가져오거나 고정값 사용)
        self.hotel_id = os.environ.get('BOOKING_HOTEL_ID', '2087141')
        
        self.driver = None
        self.wait = None
        
        logger.info(f"BookingDownloader 초기화")
        logger.info(f"  Base dir: {self.base_dir}")
        logger.info(f"  Download dir: {self.download_dir}")
        logger.info(f"  Cookies file: {self.cookies_file}")
        logger.info(f"  Hotel ID: {self.hotel_id}")
    
    def setup_driver(self):
        """Chrome WebDriver 설정"""
        logger.info("Chrome WebDriver 설정 중...")
        
        options = webdriver.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-blink-features=AutomationControlled')
        
        # 다운로드 설정
        temp_download_dir = str(self.base_dir / 'temp_downloads')
        Path(temp_download_dir).mkdir(exist_ok=True)
        
        prefs = {
            'download.default_directory': temp_download_dir,
            'download.prompt_for_download': False,
            'download.directory_upgrade': True,
            'safebrowsing.enabled': True,
            'profile.default_content_setting_values.automatic_downloads': 1
        }
        options.add_experimental_option('prefs', prefs)
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        self.wait = WebDriverWait(self.driver, 10)
        self.temp_download_dir = temp_download_dir
        
        logger.info("WebDriver 설정 완료")
    
    def load_cookies(self) -> bool:
        """저장된 쿠키 로드"""
        if self.cookies_file.exists():
            try:
                with open(self.cookies_file, 'r') as f:
                    cookies = json.load(f)
                
                # 쿠키를 추가하기 전에 도메인 페이지 열기
                self.driver.get("https://www.booking.com/sign-in")
                time.sleep(2)
                
                # 쿠키 추가
                for cookie in cookies:
                    try:
                        self.driver.add_cookie(cookie)
                    except Exception as e:
                        logger.warning(f"쿠키 추가 실패: {cookie.get('name')} - {e}")
                
                logger.info(f"쿠키 로드 완료: {len(cookies)}개")
                return True
            except Exception as e:
                logger.warning(f"쿠키 로드 실패: {e}")
                return False
        return False
    
    def save_cookies(self):
        """쿠키 저장"""
        try:
            cookies = self.driver.get_cookies()
            with open(self.cookies_file, 'w') as f:
                json.dump(cookies, f, indent=2)
            logger.info(f"쿠키 저장 완료: {len(cookies)}개")
        except Exception as e:
            logger.error(f"쿠키 저장 실패: {e}")
    
    def login(self):
        """부킹 로그인"""
        try:
            logger.info("부킹 로그인 시작...")
            
            # 숙소관리자 페이지 로그인 URL
            login_url = "https://account.booking.com/sign-in?op_token=EgVvYXV0aCLqAQoUNlo3Mm9IT2QzNk5uN3prM3BpcmgSCWF1dGhvcml6ZRoaaHR0cHM6Ly9hZG1pbi5ib29raW5nLmNvbS8qcnsidXRtX3NvdXJjZSI6ImpvaW5hcHAiLCJhdXRoX2F0dGVtcHRfaWQiOiIyZWQzYzQ4Ni0zNjQyLTQyNGItODJhNi02ZDU2YmQ0OWE3MDciLCJ1dG1fbWVkaXVtIjoiam9pbmFwcC1zZW8tcGFnZXMifTIrQm94VjhjN1dFSFdNSVUzVml5aHlOSnRsZlBlSUtiYWRYWlVqLXp0MnBuODoEUzI1NkIEY29kZSovMLbe59CHmSg6AEIcCgdqb2luYXBwEhFqb2luYXBwLXNlby1wYWdlc1iviJiquTM"
            self.driver.get(login_url)
            time.sleep(3)
            
            # 쿠키 먼저 시도
            if self.load_cookies():
                self.driver.get("https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage/payouts.html")
                time.sleep(3)
                
                # 로그인 상태 확인
                try:
                    self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'finance-payouts-payout-overview')))
                    logger.info("저장된 쿠키로 로그인 성공")
                    return True
                except:
                    logger.info("저장된 쿠키가 만료됨, 새로 로그인...")
            
            # 크리덴셜 확인
            if not self.username or not self.password:
                raise ValueError("BOOKING_USERNAME, BOOKING_PASSWORD 환경 변수 필요")
            
            # 로그인 ID 입력 (loginname 필드)
            logger.info("로그인 ID 입력 중...")
            try:
                # 먼저 loginname ID로 찾기
                loginname_input = self.wait.until(
                    EC.presence_of_element_located((By.ID, 'loginname'))
                )
                loginname_input.clear()
                loginname_input.send_keys(self.username)
                time.sleep(1)
                logger.info("로그인 ID 입력 완료")
            except:
                # 없으면 email input 사용
                logger.info("loginname 필드 없음, email 필드로 시도...")
                email_input = self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="email"]'))
                )
                email_input.clear()
                email_input.send_keys(self.username)
                time.sleep(1)
            
            # 계속 버튼 클릭
            continue_button = self.driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]')
            continue_button.click()
            logger.info("제출 완료")
            
            # 비밀번호 입력
            logger.info("비밀번호 입력 중...")
            time.sleep(5)  # 페이지 로딩 완료 대기
            
            # 비밀번호 필드 찾기
            try:
                password_input = self.wait.until(
                    EC.presence_of_element_located((By.ID, 'password'))
                )
                logger.info(f"비밀번호 필드 찾음: {password_input.tag_name}")
            except:
                logger.error("비밀번호 필드를 찾을 수 없음")
                logger.info("페이지 HTML 출력")
                logger.info(self.driver.page_source[:1000])
                raise
            
            # 비밀번호 입력 (느리게)
            password_input.clear()
            time.sleep(1)
            password_input.click()  # 필드에 focus
            time.sleep(1)
            
            # 글자 하나씩 입력
            for char in self.password:
                password_input.send_keys(char)
                time.sleep(0.05)  # 글자 하나마다 50ms 대기
            
            logger.info(f"비밀번호 입력 완료: {'*' * len(self.password)}")
            time.sleep(2)
            
            # 비밀번호 입력 확인
            current_value = password_input.get_attribute('value')
            logger.info(f"입력된 비밀번호 값 확인: {len(current_value)}자 (기대값: {len(self.password)}자)")
            
            if len(current_value) == 0:
                logger.error("비밀번호가 입력되지 않았습니다!")
                logger.info("수동 입력으로 재시도 중...")
                password_input.send_keys(self.password)
                time.sleep(2)
                current_value = password_input.get_attribute('value')
                logger.info(f"재입력 후 확인: {len(current_value)}자")
            
            # Sign in 버튼 클릭
            logger.info("Sign in 버튼 찾기...")
            try:
                # span으로 "Sign in" 텍스트 찾기
                signin_button = self.driver.find_element(By.XPATH, "//span[contains(text(), 'Sign in')]/..")
                logger.info(f"Sign in 버튼 찾음: {signin_button.tag_name}, enabled: {signin_button.is_enabled()}")
                
                # 버튼을 스크롤하여 보이게 하기
                self.driver.execute_script("arguments[0].scrollIntoView(true);", signin_button)
                time.sleep(1)
                
                # JavaScript로 클릭
                self.driver.execute_script("arguments[0].click();", signin_button)
                logger.info("Sign in 버튼 JavaScript 클릭 완료")
            except Exception as e:
                # 실패하면 submit 버튼으로 시도
                logger.warning(f"Sign in 버튼 못 찾음: {e}")
                logger.info("submit 버튼으로 시도...")
                try:
                    submit_button = self.driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]')
                    self.driver.execute_script("arguments[0].click();", submit_button)
                    logger.info("제출 버튼 JavaScript 클릭 완료")
                except Exception as e2:
                    logger.error(f"제출 버튼도 실패: {e2}")
                    raise
            
            # 로그인 완료 확인 (충분한 대기 시간 확보)
            logger.info("로그인 처리 중... (최대 30초 대기)")
            time.sleep(15)  # 더 긴 대기 시간
            
            # 현재 URL 확인 및 리다이렉트 대기
            logger.info(f"현재 URL: {self.driver.current_url}")
            
            # admin.booking.com로 리다이렉트 대기 (또는 완료 확인)
            try:
                # URL이 admin.booking.com으로 변경될 때까지 대기
                self.wait.until(lambda driver: 'admin.booking.com' in driver.current_url or 
                               driver.find_element(By.CLASS_NAME, 'finance-payouts-payout-overview'))

                logger.info(f"로그인 성공 - 현재 URL: {self.driver.current_url}")
                self.save_cookies()
                return True
            except:
                # 리다이렉트 실패 시, 수동으로 admin 페이지로 이동
                logger.warning("자동 리다이렉트 실패, 수동으로 admin 페이지로 이동...")
                current_url = self.driver.current_url
                logger.info(f"현재 URL: {current_url}")
                
                if 'admin.booking.com' not in current_url:
                    logger.info("admin.booking.com으로 이동 시도...")
                    self.driver.get("https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage/payouts.html")
                    time.sleep(5)
                    
                    try:
                        self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'finance-payouts-payout-overview')))
                        logger.info("관리자 페이지 진입 성공")
                        self.save_cookies()
                        return True
                    except:
                        logger.error("관리자 페이지 진입 실패")
                        return False
                else:
                    logger.error("로그인 완료 확인 실패")
                    return False
        
        except Exception as e:
            logger.error(f"로그인 실패: {e}")
            return False
    
    def navigate_to_payouts(self):
        """대금지급정보 페이지로 이동"""
        try:
            logger.info("대금지급정보 페이지 이동 중...")
            
            payouts_url = f"https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage/payouts.html?hotel_id={self.hotel_id}&lang=ko"
            self.driver.get(payouts_url)
            time.sleep(3)
            
            # 명세서 테이블 로드 대기
            self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'finance-payouts-payout-overview')))
            logger.info("대금지급정보 페이지 로드 완료")
            return True
        
        except Exception as e:
            logger.error(f"페이지 이동 실패: {e}")
            return False
    
    def parse_payouts_table(self) -> List[Dict]:
        """명세서 테이블 파싱"""
        try:
            logger.info("명세서 테이블 파싱 중...")
            
            # 테이블 행 찾기
            rows = self.driver.find_elements(By.CSS_SELECTOR, 'tbody.finance-payouts-payout-overview tr')
            logger.info(f"발견된 행: {len(rows)}개")
            
            payouts = []
            
            for idx, row in enumerate(rows):
                try:
                    cells = row.find_elements(By.TAG_NAME, 'td')
                    
                    if len(cells) < 7:
                        continue
                    
                    # 지급일 (2번째 셀)
                    payout_date = cells[1].text.strip()  # "2026년 1월 1일"
                    
                    # 기간 (3번째 셀)
                    period_text = cells[2].text.strip()  # "12월 1일 ~ 12월 31일"
                    
                    # 대금지급ID (4번째 셀)
                    payout_id = cells[3].text.strip()
                    
                    # 금액 (마지막 셀)
                    amount_text = cells[-1].text.strip()  # "₩54,895,500"
                    
                    # 날짜 파싱
                    try:
                        match = re.match(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', payout_date)
                        if match:
                            year, month, day = match.groups()
                            payout_date_obj = datetime(int(year), int(month), int(day))
                            payout_date_str = payout_date_obj.strftime('%Y-%m-%d')
                        else:
                            continue
                    except:
                        continue
                    
                    # 기간 파싱 ("12월 1일 ~ 12월 31일" → 12-01, 12-31)
                    period_from = None
                    period_to = None
                    try:
                        matches = re.findall(r'(\d{1,2})월\s*(\d{1,2})일', period_text)
                        if len(matches) >= 2:
                            month_from, day_from = matches[0]
                            month_to, day_to = matches[1]
                            period_from = f"{int(month_from):02d}-{int(day_from):02d}"
                            period_to = f"{int(month_to):02d}-{int(day_to):02d}"
                    except:
                        pass
                    
                    # 금액 파싱 ("₩54,895,500" → 54895500)
                    try:
                        amount_str = re.sub(r'[^\d.]', '', amount_text)
                        amount_int = int(float(amount_str)) if amount_str else 0
                    except:
                        amount_int = 0
                    
                    if payout_id and payout_date_str:
                        payout = {
                            'payout_id': payout_id,
                            'payout_date': payout_date_str,
                            'period_from': period_from,
                            'period_to': period_to,
                            'amount': amount_int,
                            'download_url': None
                        }
                        payouts.append(payout)
                
                except Exception as e:
                    logger.warning(f"행 파싱 실패 (행 {idx}): {e}")
                    continue
            
            logger.info(f"파싱 완료: {len(payouts)}개 명세서")
            return payouts
        
        except Exception as e:
            logger.error(f"테이블 파싱 실패: {e}")
            return []
    
    def download_payout_csv(self, payout_id: str, amount: int) -> bool:
        """개별 명세서 CSV 다운로드"""
        try:
            logger.info(f"다운로드: {payout_id} ({amount:,} KRW)")
            
            # 파일명 생성: 부킹_YYYYMMDD_금액.csv
            date_obj = datetime.now()
            date_str = date_obj.strftime('%Y%m%d')
            target_filename = f"부킹_{date_str}_{amount}.csv"
            target_file = self.download_dir / target_filename
            
            # 이미 존재하면 스킵
            if target_file.exists():
                logger.info(f"파일 이미 존재: {target_filename}")
                return False
            
            # 행 찾기 및 다운로드 링크 클릭
            try:
                # 해당 행 찾기
                rows = self.driver.find_elements(By.CSS_SELECTOR, 'tbody.finance-payouts-payout-overview tr')
                for row in rows:
                    if payout_id in row.text:
                        # CSV 다운로드 링크 찾기
                        csv_link = row.find_element(By.XPATH, './/a[contains(text(), "CSV")]')
                        
                        # 다운로드 링크 클릭
                        self.driver.execute_script("arguments[0].click();", csv_link)
                        time.sleep(2)
                        logger.info(f"CSV 다운로드 클릭 완료")
                        
                        # 다운로드 파일 확인
                        temp_dir = Path(self.temp_download_dir)
                        files = sorted([p for p in temp_dir.glob('*.csv')], 
                                      key=lambda p: p.stat().st_mtime, 
                                      reverse=True)
                        
                        if files:
                            source_file = files[0]
                            source_file.rename(target_file)
                            logger.info(f"파일 저장: {target_filename}")
                            return True
                        else:
                            logger.warning(f"다운로드 파일을 찾을 수 없음")
                            return False
            except NoSuchElementException:
                logger.warning(f"CSV 링크를 찾을 수 없음: {payout_id}")
                return False
        
        except Exception as e:
            logger.error(f"다운로드 실패: {payout_id} - {e}")
            return False
    
    def download_payouts(self) -> List[Dict]:
        """명세서 일괄 다운로드"""
        try:
            logger.info("명세서 일괄 다운로드 시작...")
            
            # 테이블 파싱
            payouts = self.parse_payouts_table()
            
            if not payouts:
                logger.error("다운로드할 명세서가 없습니다")
                return []
            
            # 개별 다운로드
            downloaded_count = 0
            for payout in payouts:
                if self.download_payout_csv(payout['payout_id'], payout['amount']):
                    downloaded_count += 1
                time.sleep(1)
            
            logger.info(f"다운로드 완료: {downloaded_count}개")
            return payouts
        
        except Exception as e:
            logger.error(f"명세서 다운로드 실패: {e}")
            return []
    
    def update_excel_with_payouts(self, payouts: List[Dict]):
        """엑셀 파일에 명세서 추가"""
        try:
            excel_path = self.base_dir / '매출 및 입금 결과.xlsx'
            
            logger.info(f"엑셀 업데이트: {excel_path}")
            
            # 파일이 없으면 생성
            if not excel_path.exists():
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            else:
                wb = load_workbook(excel_path)
            
            # 부킹 시트 생성 또는 선택
            if '부킹' not in wb.sheetnames:
                ws = wb.create_sheet('부킹')
                ws.append(['대금지급기간', '시작일', '종료일', '대금'])
                logger.info("'부킹' 시트 생성")
            else:
                ws = wb['부킹']
            
            # 기존 데이터 읽기
            existing_payout_ids = set()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    existing_payout_ids.add(str(row[0]).strip())
            
            logger.info(f"기존 데이터: {len(existing_payout_ids)}건")
            
            # 새로운 명세서 추가
            added_count = 0
            for payout in payouts:
                if payout['payout_id'] in existing_payout_ids:
                    continue
                
                amount_formatted = f"{payout['amount']:,}"
                
                ws.append([
                    payout['payout_id'],
                    payout['period_from'],
                    payout['period_to'],
                    amount_formatted
                ])
                
                existing_payout_ids.add(payout['payout_id'])
                added_count += 1
                logger.info(f"추가: {payout['payout_id']} - {amount_formatted} KRW")
            
            logger.info(f"추가된 명세서: {added_count}개")
            
            # 정렬
            if ws.max_row > 1:
                data_rows = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    data_rows.append(list(row))
                
                data_rows.sort(key=lambda x: str(x[0]) if x[0] else '', reverse=True)
                
                ws.delete_rows(2, ws.max_row - 1)
                
                for row_data in data_rows:
                    ws.append(row_data)
                
                logger.info("정렬 완료")
            
            # 저장
            wb.save(excel_path)
            wb.close()
            logger.info("엑셀 저장 완료")
        
        except Exception as e:
            logger.error(f"엑셀 업데이트 실패: {e}")
    
    def close(self):
        """WebDriver 종료"""
        if self.driver:
            self.driver.quit()
            logger.info("WebDriver 종료")
    
    def run(self):
        """전체 실행"""
        try:
            self.setup_driver()
            
            if not self.login():
                logger.error("로그인 실패")
                print("\n[ERROR] 로그인 실패했습니다.")
                print("브라우저 창을 확인하고 Enter를 눌러주세요...")
                input()
                return
            
            if not self.navigate_to_payouts():
                logger.error("페이지 이동 실패")
                print("\n[ERROR] 페이지 이동 실패했습니다.")
                print("브라우저 창을 확인하고 Enter를 눌러주세요...")
                input()
                return
            
            payouts = self.download_payouts()
            
            if payouts:
                self.update_excel_with_payouts(payouts)
            
            logger.info("부킹닷컴 명세서 다운로드 완료")
            print("\n[완료] 부킹닷컴 명세서 다운로드 완료")
        
        except Exception as e:
            logger.error(f"실행 중 오류: {e}")
            print(f"\n[ERROR] 오류 발생: {e}")
            print("브라우저 창을 확인하고 Enter를 눌러주세요...")
            import traceback
            traceback.print_exc()
            input()
        finally:
            # 사용자가 창을 닫도록 선택하지 않은 경우에만 종료
            keep_open = os.environ.get('BOOKING_KEEP_BROWSER_OPEN', '0')
            if keep_open.lower() not in ('1', 'true', 'yes'):
                # 정상 완료 시만 자동으로 종료
                try:
                    # 오류가 발생하지 않았으면 3초 후 자동 종료
                    time.sleep(3)
                    self.close()
                except:
                    pass
            else:
                logger.info("BOOKING_KEEP_BROWSER_OPEN 설정으로 브라우저를 열어둡니다.")


if __name__ == '__main__':
    """명령행 실행"""
    downloader = BookingDownloader()
    downloader.run()
    
    def detect_booking_files(self) -> List[Tuple[Path, str]]:
        """
        부킹 CSV 파일 감지 및 정규화
        
        Returns:
            [(파일경로, 정규화된파일명), ...] 리스트
        """
        print("\n[파일 감지] 부킹 CSV 파일 검색 중...")
        
        booking_files = []
        
        # 1. 이미 부킹_으로 시작하는 파일
        for f in self.download_dir.glob('부킹_*.csv'):
            booking_files.append((f, f.name))
        
        # 2. 부킹 명세서인데 정규화되지 않은 파일
        # 부킹 CSV는 보통 "Payout-*.csv" 형식일 수 있음
        for f in self.download_dir.glob('*.csv'):
            name = f.name
            # 이미 처리된 파일은 스킵
            if name.startswith('부킹_') or name.startswith('아고다_') or name.startswith('익스피디아_'):
                continue
            # 부킹 파일 가능성 판단 (부킹 CSV 헤더 확인)
            try:
                df = pd.read_csv(f, nrows=1, encoding='utf-8-sig')
                # 부킹 CSV 특성: "Payout Status", "Payout Amount" 등의 컬럼
                cols_lower = [str(c).lower() for c in df.columns]
                if any('payout' in c or 'amount' in c for c in cols_lower):
                    print(f"  [발견] 부킹 파일로 추정: {name}")
                    booking_files.append((f, name))
            except Exception as e:
                # CSV 읽기 실패하면 스킵
                pass
        
        print(f"  발견된 부킹 파일: {len(booking_files)}개")
        return booking_files
    
    def parse_booking_csv(self, file_path: Path) -> List[dict]:
        """
        부킹 CSV 파일 파싱
        
        Args:
            file_path: CSV 파일 경로
            
        Returns:
            명세서 레코드 리스트
        """
        try:
            print(f"\n[파싱] {file_path.name}")
            
            # CSV 읽기 (부킹 CSV는 보통 UTF-8-sig 인코딩)
            df = pd.read_csv(file_path, encoding='utf-8-sig')
            
            print(f"  컬럼: {list(df.columns[:10])}")
            print(f"  행: {len(df)}개")
            
            records = []
            
            for idx, row in df.iterrows():
                try:
                    # 부킹 CSV 구조:
                    # - Payout Date (또는 "Payout issued date"): 지급일 (YYYY년 M월 D일 형식)
                    # - Period from / Period to: 기간 (M월 D일 형식)
                    # - Payout ID (또는 "Payout UUID"): 지급 ID
                    # - Amount (또는 "Payout Amount"): 금액
                    
                    # 컬럼 찾기 (한글/영문 모두 대응)
                    payout_date = None
                    payout_period_from = None
                    payout_period_to = None
                    payout_id = None
                    amount = None
                    
                    for col in df.columns:
                        col_lower = str(col).lower()
                        
                        if 'date' in col_lower or '날짜' in col:
                            payout_date = row[col]
                        elif 'period' in col_lower and ('from' in col_lower or '시작' in col):
                            payout_period_from = row[col]
                        elif 'period' in col_lower and ('to' in col_lower or '종료' in col):
                            payout_period_to = row[col]
                        elif 'id' in col_lower or '아이디' in col or 'uuid' in col_lower:
                            payout_id = row[col]
                        elif 'amount' in col_lower or '금액' in col:
                            amount = row[col]
                    
                    # 필수 필드 확인
                    if not payout_id or not amount:
                        continue
                    
                    # 날짜 파싱 (한글 형식: "2026년 1월 1일" → YYYY-MM-DD)
                    try:
                        if payout_date:
                            # "2026년 1월 1일" 형식 파싱
                            payout_date_str = str(payout_date).strip()
                            match = re.match(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', payout_date_str)
                            if match:
                                year, month, day = match.groups()
                                payout_date_obj = datetime(int(year), int(month), int(day))
                            else:
                                # 다른 형식 시도
                                payout_date_obj = pd.to_datetime(payout_date)
                        else:
                            payout_date_obj = None
                    except Exception as e:
                        print(f"    [WARN] 날짜 파싱 실패: {payout_date} - {e}")
                        payout_date_obj = None
                    
                    # 기간 파싱 ("12월 1일 ~ 12월 31일" → start_month_day, end_month_day)
                    period_from_str = None
                    period_to_str = None
                    try:
                        if payout_period_from:
                            # "12월 1일" 형식 파싱
                            period_from_str = str(payout_period_from).strip()
                            match = re.match(r'(\d{1,2})월\s*(\d{1,2})일', period_from_str)
                            if match:
                                month, day = match.groups()
                                period_from_str = f"{int(month):02d}-{int(day):02d}"
                        
                        if payout_period_to:
                            # "12월 31일" 형식 파싱
                            period_to_str = str(payout_period_to).strip()
                            match = re.match(r'(\d{1,2})월\s*(\d{1,2})일', period_to_str)
                            if match:
                                month, day = match.groups()
                                period_to_str = f"{int(month):02d}-{int(day):02d}"
                    except Exception as e:
                        print(f"    [WARN] 기간 파싱 실패: {payout_period_from} ~ {payout_period_to} - {e}")
                    
                    # 금액 파싱 (한글 통화기호 제거, 천단위 쉼표 제거)
                    try:
                        amount_str = str(amount).strip()
                        # "₩54,895,500" → "54895500"
                        amount_str = re.sub(r'[^\d.]', '', amount_str)
                        amount_float = float(amount_str) if amount_str else 0.0
                        amount_int = int(amount_float)
                    except Exception as e:
                        print(f"    [WARN] 금액 파싱 실패: {amount} - {e}")
                        amount_int = 0
                    
                    record = {
                        'payout_id': str(payout_id).strip(),
                        'payout_date': payout_date_obj,
                        'payout_date_str': payout_date_obj.strftime('%Y-%m-%d') if payout_date_obj else '',
                        'period_from': period_from_str,
                        'period_to': period_to_str,
                        'amount': amount_int,
                    }
                    
                    if record['payout_id'] and record['payout_date_str']:
                        records.append(record)
                
                except Exception as e:
                    print(f"    [WARN] 행 파싱 실패 (행 {idx}): {e}")
                    continue
            
            print(f"  파싱 완료: {len(records)}개 명세서")
            return records
        
        except Exception as e:
            print(f"  [ERROR] CSV 파싱 실패: {e}")
            return []
    
    def update_excel_with_bookings(self, records: List[dict]):
        """
        엑셀 파일에 부킹 명세서 추가
        
        Args:
            records: 명세서 레코드 리스트
        """
        try:
            excel_path = self.base_dir / '매출 및 입금 결과.xlsx'
            
            print(f"\n[엑셀 업데이트] {excel_path}")
            
            # 파일이 없으면 생성
            if not excel_path.exists():
                print(f"  엑셀 파일이 없어 새로 생성합니다.")
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            else:
                wb = load_workbook(excel_path)
            
            # 부킹 시트 생성 또는 선택
            if '부킹' not in wb.sheetnames:
                ws = wb.create_sheet('부킹')
                # 헤더 추가
                ws.append(['대금지급기간', '시작일', '종료일', '대금'])
                print(f"  '부킹' 시트 생성 및 헤더 추가")
            else:
                ws = wb['부킹']
            
            # 기존 데이터 읽기 (중복 제거용)
            existing_payout_ids = set()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    existing_payout_ids.add(str(row[0]).strip())
            
            print(f"  기존 데이터: {len(existing_payout_ids)}건")
            
            # 새로운 명세서만 추가
            added_count = 0
            for record in records:
                payout_id = record['payout_id']
                
                if payout_id in existing_payout_ids:
                    continue
                
                # 금액을 천단위 쉼표로 포맷
                amount_formatted = f"{record['amount']:,}"
                
                # 행 추가
                ws.append([
                    payout_id,
                    record['period_from'],
                    record['period_to'],
                    amount_formatted
                ])
                
                existing_payout_ids.add(payout_id)
                added_count += 1
                
                print(f"  추가: {payout_id} - {amount_formatted} KRW")
            
            print(f"  추가된 명세서: {added_count}개")
            
            # 지급일 기준 정렬 (최근순)
            if ws.max_row > 1:
                data_rows = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    data_rows.append(list(row))
                
                # 대금지급기간ID(A열)로 정렬 (역순이므로 내림차순)
                # 실제로는 지급일 기준 정렬을 원하면 record의 payout_date 사용
                # 현재는 ID 기준 정렬 (대금지급기간ID가 타임스탬프 포함 형식이므로 적당함)
                data_rows.sort(key=lambda x: str(x[0]) if x[0] else '', reverse=True)
                
                # 기존 데이터 삭제
                ws.delete_rows(2, ws.max_row - 1)
                
                # 정렬된 데이터 다시 추가
                for row_data in data_rows:
                    ws.append(row_data)
                
                print(f"  정렬 완료: 대금지급기간 기준 정렬")
            
            # 파일 저장
            wb.save(excel_path)
            wb.close()
            print(f"  [OK] 엑셀 저장 완료")
        
        except Exception as e:
            print(f"  [ERROR] 엑셀 업데이트 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def run(self):
        """전체 실행 (파일 감지 → 파싱 → 엑셀 업데이트)"""
        try:
            # 부킹 파일 감지
            booking_files = self.detect_booking_files()
            
            if not booking_files:
                print("\n[결과] 부킹 CSV 파일을 찾을 수 없습니다.")
                print("  다운로드 경로: https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage/payouts.html")
                print("  '모든 명세서 다운로드' → CSV 파일 → ota-adjustment 폴더에 저장")
                return
            
            # 모든 파일 파싱
            all_records = []
            for file_path, file_name in booking_files:
                records = self.parse_booking_csv(file_path)
                all_records.extend(records)
            
            if not all_records:
                print("\n[결과] 파싱된 명세서가 없습니다.")
                return
            
            print(f"\n[통계] 총 {len(all_records)}개 명세서 파싱 완료")
            
            # 엑셀 업데이트
            self.update_excel_with_bookings(all_records)
            
            print("\n[완료] 부킹닷컴 명세서 처리 완료")
        
        except Exception as e:
            print(f"[ERROR] 실행 중 오류: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    """명령행 실행"""
    downloader = BookingDownloader()
    downloader.run()
