# -*- coding: utf-8 -*-
import sys
import io
import argparse

# Windows 콘솔 인코딩 문제 해결
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 명령행 인자 파싱
parser = argparse.ArgumentParser(description='호텔 매출 비교 프로그램')
parser.add_argument('--download-expedia', action='store_true', help='Expedia 명세서 자동 다운로드 실행')
parser.add_argument('--expedia-start-date', help='Expedia 다운로드 시작 날짜 (YYYY-MM-DD)')
parser.add_argument('--expedia-end-date', help='Expedia 다운로드 종료 날짜 (YYYY-MM-DD)')
args = parser.parse_args()

# Expedia 다운로드 옵션 처리
if args.download_expedia:
    print("\n" + "="*80)
    print("Expedia 명세서 다운로드 옵션 활성화")
    print("="*80)
    try:
        from expedia_downloader import ExpediaDownloader
        
        downloader = ExpediaDownloader(base_dir=os.path.dirname(os.path.abspath(__file__)))
        count = downloader.run(
            start_date=args.expedia_start_date,
            end_date=args.expedia_end_date
        )
        
        print(f"\n[결과] {count}개 Expedia 명세서 다운로드 완료\n")
    except Exception as e:
        print(f"\n[ERROR] Expedia 다운로드 실패: {e}\n")
        import traceback
        traceback.print_exc()

# ...기존 코드...

# 파일 맨 아래에 print_result_rows 정의 및 호출

# ...기존 코드 맨 아래에 추가...
def write_ratio_to_result_log(ratio, row_idx):
    from openpyxl import load_workbook
    result_path = os.path.join(dir_base, '매출_검토_결과.xlsx')
    wb = load_workbook(result_path)
    if '비교로그' not in wb.sheetnames:
        print('[비교로그] 시트가 없습니다.')
        return
    ws = wb['비교로그']
    # row_idx는 1부터 시작(헤더 포함), G열(7번째 열)에 기록
    ws.cell(row=row_idx, column=7).value = ratio
    wb.save(result_path)
    print(f"[진단-비교로그] {row_idx}행 G열에 비율 {ratio} 기록 완료")
import re

def normalize(val):
    if val is None:
        return ''
    return re.sub(r'[^a-zA-Z0-9가-힣]', '', str(val)).replace('.0','').strip().lower()
def print_peter_ludwig_log():
    from openpyxl import load_workbook
    result_path = os.path.join(dir_base, '매출_검토_결과.xlsx')
    wb = load_workbook(result_path, data_only=True)
    if '비교로그' not in wb.sheetnames:
        print('[비교로그] 시트가 없습니다.')
        return
    ws = wb['비교로그']
    # print('[비교로그] Peter Ludwig 관련 행:')
    for row in ws.iter_rows(values_only=True):
        if row and any('Peter Ludwig' in str(cell) for cell in row):
            print(row)

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from collections import defaultdict
import re

# 파일 경로 설정
dir_base = os.path.dirname(os.path.abspath(__file__))

directory_ota = os.path.join(dir_base, 'ota-adjustment')

# 결과 엑셀 파일에서만 데이터 로드

# 결과파일이 없으면 전체고객 목록 파일을 복사해서 생성
result_path = os.path.join(dir_base, '매출_검토_결과.xlsx')
if not os.path.exists(result_path):
    import glob
    import shutil
    all_list = sorted(glob.glob(os.path.join(dir_base, '전체고객 목록_*.xlsx')))
    if not all_list:
        raise FileNotFoundError('전체고객 목록_*.xlsx 파일이 없습니다. 결과파일을 생성할 수 없습니다.')
    latest_all = all_list[-1]
    shutil.copy(latest_all, result_path)
df_all = pd.read_excel(result_path, sheet_name=0)

# 컬럼명 매핑 (자동 추출)
def find_col(cols, keyword):
    for c in cols:
        if keyword in c:
            return c
    return None

# OTA번호 컬럼이 있으면 항상 문자열로 변환(.0 제거)
col_ota_no = find_col(df_all.columns, 'OTA')
if col_ota_no:
    df_all[col_ota_no] = df_all[col_ota_no].astype(str).str.replace('.0', '', regex=False).str.strip()

# Remittances로 시작하는 파일 목록
ota_files = [f for f in os.listdir(directory_ota) if f.startswith('Remittances') and f.endswith('.xlsx')]

# 아고다 매출 데이터 통합
df_ota = pd.DataFrame()
for file in ota_files:
    path = os.path.join(directory_ota, file)
    temp_df = pd.read_excel(path)
    df_ota = pd.concat([df_ota, temp_df], ignore_index=True)

# 부킹 CSV 파일 읽기
booking_files = [f for f in os.listdir(directory_ota) if f.startswith('부킹') and f.endswith('.csv')]
df_booking = pd.DataFrame()
for file in booking_files:
    path = os.path.join(directory_ota, file)
    temp_df = pd.read_csv(path)
    df_booking = pd.concat([df_booking, temp_df], ignore_index=True)

# 부킹 데이터 구조: B열=예약번호, I열=가격
if not df_booking.empty:
    # 예약번호를 문자열로 변환 (FutureWarning 방지)
    df_booking[df_booking.columns[1]] = df_booking.iloc[:, 1].astype(str).str.strip()
    # 금액 컬럼 (I열 = 인덱스 8)
    booking_price_col = df_booking.columns[8] if len(df_booking.columns) > 8 else None

# 익스피디아 CSV 파일 읽기
expedia_files = [f for f in os.listdir(directory_ota) if f.startswith('익스피디아') and f.endswith('.csv')]
df_expedia = pd.DataFrame()
for file in expedia_files:
    path = os.path.join(directory_ota, file)
    temp_df = pd.read_csv(path)
    df_expedia = pd.concat([df_expedia, temp_df], ignore_index=True)

# 익스피디아 데이터 구조: A열=예약번호, F열=처리금액
if not df_expedia.empty:
    # 예약번호를 문자열로 변환 (FutureWarning 방지)
    df_expedia[df_expedia.columns[0]] = df_expedia.iloc[:, 0].astype(str).str.strip()
    # 금액 컬럼 (F열 = 인덱스 5)
    expedia_price_col = df_expedia.columns[5] if len(df_expedia.columns) > 5 else None




# 컬럼명 매핑 (자동 추출)
def find_col(cols, keyword):
    for c in cols:
        if keyword in c:
            return c
    return None

col_name_all = find_col(df_all.columns, '고객')
col_price_all_1 = find_col(df_all.columns, '객실')
col_price_all_2 = find_col(df_all.columns, '합계')
col_vendor = find_col(df_all.columns, '거래처')
col_ota_no = find_col(df_all.columns, 'OTA')

"""
# print(f"[DEBUG] 컬럼명 매핑: 고객명={col_name_all}, 객실료={col_price_all_1}, 합계={col_price_all_2}, 거래처={col_vendor}, OTA번호={col_ota_no}")
"""

# Remittances 엑셀의 이름, 금액 컬럼명 추정 (수정 필요시 아래 변수명 변경)
# 이름 컬럼: 4번째 컬럼(D열)
col_name_ota = df_ota.columns[3]
# 금액 컬럼: Remittances의 G열, H열 등 여러 컬럼을 모두 비교
ota_price_cols = [col for col in df_ota.columns if any(x in col for x in ['금액', 'Amount', '금액', '금액', '금액']) or col in ['G', 'H'] or col.startswith('Unnamed')]
if not ota_price_cols:
    # Remittances의 7,8번째 컬럼(G,H열)로 강제 지정
    ota_price_cols = [df_ota.columns[6], df_ota.columns[7]]

result_path = os.path.join(dir_base, '매출_검토_결과.xlsx')
wb = load_workbook(result_path)
ws = wb.active
# 비교로그 시트 생성(기존 있으면 삭제)
if '비교로그' in wb.sheetnames:
    del wb['비교로그']
log_ws = wb.create_sheet('비교로그')
log_ws.append(['고객명', '전체매출 행번호', '전체매출 가격', '파일명', '행번호', '비교 가격', '원가격'])

# Remittances 파일별 행 오프셋 기록비교 가격
ota_file_map = []  # (파일명, 데이터프레임, 시작행)
ota_row_offset = 0
df_ota = pd.DataFrame()
for file in ota_files:
    path = os.path.join(directory_ota, file)
    temp_df = pd.read_excel(path)
    ota_file_map.append((file, temp_df, ota_row_offset))
    ota_row_offset += len(temp_df)
    df_ota = pd.concat([df_ota, temp_df], ignore_index=True)

# 색상 스타일 정의
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_blue = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
font_red = Font(color='FF0000')


# 이름과 가격 비교 및 색상 처리
# Remittances 이름 매칭 카운트 추적용
matched_remit_names = {}

# 1. 전체고객목록에서 아고다인 고객명별로 인덱스와 가격(합계, 객실료) 수집
agoda_grouped_rows = defaultdict(list)
for idx, row in df_all.iterrows():
    vendor = str(row.get('거래처', '')).strip()
    if vendor != '아고다':
        continue
    name = str(row.get(col_name_all, '')).strip()
    price1 = str(row.get(col_price_all_1, '')).replace(',', '').strip()
    price2 = str(row.get(col_price_all_2, '')).replace(',', '').strip()
    try:
        price1_f = float(price1)
    except:
        price1_f = 0.0
    try:
        price2_f = float(price2)
    except:
        price2_f = 0.0
    # 합계 우선, 없으면 객실료
    use_price = price2_f if price2_f else price1_f
    agoda_grouped_rows[name].append((idx, use_price))

# 2. Remittances에서 이름별 금액 리스트 수집
otas_by_name = defaultdict(list)
for _, row in df_ota.iterrows():
    name = str(row[col_name_ota]).strip()
    for price_col in ota_price_cols:
        try:
            price = float(str(row[price_col]).replace(',', '').strip())
            otas_by_name[name].append(price)
        except:
            continue

# 3. 매칭 처리
used_ota_idx = set()
for name, rows in agoda_grouped_rows.items():
    total_price = sum(price for _, price in rows)
    # Remittances에서 해당 이름의 금액 중 합과 일치하는 것 찾기
    found = False
    for i, price in enumerate(otas_by_name.get(name, [])):
        if price == total_price and (name, i) not in used_ota_idx:
            # 매칭된 Remittances 금액은 중복 사용 방지
            used_ota_idx.add((name, i))
            found = True
            break
    if found:
        # 전체고객목록의 해당 이름 모든 행을 노란색으로 표시
        for idx, _ in rows:
            for cell in ws[idx+2]:
                cell.fill = fill_yellow
    else:
        # 기존 개별 비교 로직(잔여 케이스) 수행
        for idx, _ in rows:
            row = df_all.iloc[idx]
            vendor = str(row.get('거래처', '')).strip()
            if vendor != '아고다':
                continue
            name = str(row.get(col_name_all, '')).strip()
            price1 = str(row.get(col_price_all_1, '')).replace(',', '').strip()
            price2 = str(row.get(col_price_all_2, '')).replace(',', '').strip()
            # 금액을 float으로 변환
            try:
                price1_f = float(price1)
            except:
                price1_f = None
            try:
                price2_f = float(price2)
            except:
                price2_f = None
            # 합계 우선, 없으면 객실료
            use_price = price2_f if price2_f else price1_f
            # 이름이 일치하는 Remittances 행 찾기
            match = df_ota[df_ota[col_name_ota].astype(str).str.strip() == name]
            if not match.empty:
                price_match = False
                log_info = None
                for abs_idx, match_row in match.iterrows():
                    for price_col in ota_price_cols:
                        try:
                            ota_price_f = float(str(match_row[price_col]).replace(',', '').strip())
                        except:
                            continue
                        if (price1_f is not None and ota_price_f == price1_f) or (price2_f is not None and ota_price_f == price2_f):
                            price_match = True
                            # Remittances 매칭 카운트 기록
                            matched_remit_names[name] = matched_remit_names.get(name, 0) + 1
                            break
                        # 로그 정보 저장 (조건 불일치 시)
                        if not price_match and log_info is None:
                            for fname, df, offset in ota_file_map:
                                if offset <= abs_idx < offset + len(df):
                                    file_row = abs_idx - offset + 2  # 2: 엑셀 헤더 보정
                                    log_info = [name, idx+2, use_price, fname, file_row, str(match_row[price_col]), str(match_row[price_col])]
                                    break
                    if price_match:
                        break
                if price_match:
                    for cell in ws[idx+2]:
                        cell.fill = fill_yellow
                else:
                    # Remittances에 이름이 있지만, 이미 매칭된 횟수 이상이면 표시 없음
                    if matched_remit_names.get(name, 0) > 0:
                        matched_remit_names[name] -= 1
                        continue
                    for cell in ws[idx+2]:
                        cell.font = font_red
                    if log_info:
                        log_ws.append(log_info)
                    else:
                        # log_info가 None인 경우에도 비교로그에 한 줄 남김
                        log_ws.append([name, idx+2, use_price, '-', '-', '불일치', '-'])
            else:
                # Remittances에 이름이 없음 -> 파란색
                for cell in ws[idx+2]:
                    cell.fill = fill_blue
                # 비교로그에 기록
                log_ws.append([name, idx+2, use_price, '아고다 데이터 없음', '', '', ''])


# 부킹닷컴 비교 처리
print("\n" + "="*80)
print("부킹닷컴 비교 시작")
print("="*80)
matched_booking_refs = {}

# 1단계: 부킹닷컴 예약번호별 그룹화 (앞 10자리 기준)
booking_grouped_by_ref = defaultdict(list)
booking_grouped_rows = defaultdict(list)
for idx, row in df_all.iterrows():
    vendor = str(row.get('거래처', '')).strip()
    if vendor != '부킹닷컴':
        continue
    name = str(row.get(col_name_all, '')).strip()
    ota_no = str(row.get(col_ota_no, '')).strip()[:10]
    price1 = str(row.get(col_price_all_1, '')).replace(',', '').strip()
    price2 = str(row.get(col_price_all_2, '')).replace(',', '').strip()
    try:
        price1_f = float(price1)
    except:
        price1_f = 0.0
    try:
        price2_f = float(price2)
    except:
        price2_f = 0.0
    use_price = price2_f if price2_f else price1_f
    
    # 예약번호별 그룹화
    if ota_no:
        booking_grouped_by_ref[ota_no].append((idx, use_price, name))
    # 고객명별 그룹화 (백업용)
    booking_grouped_rows[name].append((idx, use_price))

print(f"\n[1단계] 전체고객목록에서 부킹닷컴 예약번호 {len(booking_grouped_by_ref)}개, 고객 {len(booking_grouped_rows)}명 그룹화 완료")

# 2단계: 부킹 데이터 수집
booking_by_ref = {}
if not df_booking.empty:
    print(f"\n[2단계] 부킹 CSV 파일 데이터 읽기 시작 (총 {len(df_booking)}행)")
    print(f"부킹 CSV 컬럼: {list(df_booking.columns[:10])}")
    
    for b_idx, b_row in df_booking.iterrows():
        try:
            b_ref = str(b_row.iloc[1]).strip()
            if booking_price_col and booking_price_col in df_booking.columns:
                b_price = float(str(b_row[booking_price_col]).replace(',', '').strip())
            else:
                b_price = float(str(b_row.iloc[8]).replace(',', '').strip())
            booking_by_ref[b_ref] = round(b_price * 0.82)
        except:
            continue

print("\n[3단계] 예약번호 기준 그룹 합산 매칭 시작")
used_booking_refs = set()
group_matched_count = 0
matched_rows = set()

for ref_no, rows in booking_grouped_by_ref.items():
    total_price = sum(price for _, price, _ in rows)
    found = False
    
    customer_names = ', '.join(set(name for _, _, name in rows))
    print(f"\n예약번호: {ref_no} (고객명: {customer_names})")
    print(f"  전체고객목록 행 수: {len(rows)}, 가격 합계: {total_price}")
    print(f"  부킹 데이터 가격: {booking_by_ref.get(ref_no, 'N/A')}")
    print(f"  부킹 데이터 가격: {booking_by_ref.get(ref_no, 'N/A')}")
    
    if ref_no in booking_by_ref and round(total_price) == booking_by_ref[ref_no]:
        found = True
        used_booking_refs.add(ref_no)
        print(f"  [OK] 예약번호 그룹 합산 매칭 성공! (전체고객목록 합계: {total_price} = 부킹 가격: {booking_by_ref[ref_no]})")
    
    if found:
        group_matched_count += 1
        for idx, _, _ in rows:
            matched_rows.add(idx)
            for cell in ws[idx+2]:
                cell.fill = fill_yellow
        print(f"  → {len(rows)}개 행 모두 노란색 표시")
    else:
        print(f"  [SKIP] 예약번호 그룹 합산 매칭 실패")

print(f"\n[완료] 예약번호 기준 그룹 합산 매칭: {group_matched_count}건, {len(matched_rows)}개 행 처리됨")

# 4단계: 매칭되지 않은 행에 대해 개별 행 매칭
print("\n[4단계] 개별 행 매칭 시작 (그룹 합산 실패한 행만)")
for name, rows in booking_grouped_rows.items():
    for idx, _ in rows:
        if idx in matched_rows:
            continue
        
        row = df_all.iloc[idx]
        vendor = str(row.get('거래처', '')).strip()
        if vendor != '부킹닷컴':
            continue
        
        ws_row = idx + 2
        name = str(row.get(col_name_all, '')).strip()
        ota_no = str(row.get(col_ota_no, '')).strip()[:10]
        price1 = str(row.get(col_price_all_1, '')).replace(',', '').strip()
        price2 = str(row.get(col_price_all_2, '')).replace(',', '').strip()
        
        try:
            price1_f = float(price1)
        except:
            price1_f = None
        try:
            price2_f = float(price2)
        except:
            price2_f = None
        
        use_price = price2_f if price2_f else price1_f
        if use_price is None:
            continue
        
        if df_booking.empty:
            continue
        
        booking_match = df_booking[df_booking[df_booking.columns[1]] == ota_no]
        
        print(f"  행 {ws_row}: OTA번호={ota_no}, 가격={use_price}, 부킹매칭={len(booking_match)}건")
        
        if booking_match.empty:
            print(f"    → 부킹 데이터에 예약번호 없음 (파란색 표시)")
            for cell in ws[ws_row]:
                cell.fill = fill_blue
            # 비교로그에 기록
            log_ws.append([name, ws_row, use_price, '부킹닷컴 데이터 없음', '', '', ''])
            continue
        
        price_match = False
        log_info = None
        
        for b_idx, b_row in booking_match.iterrows():
            try:
                if booking_price_col and booking_price_col in df_booking.columns:
                    booking_price = float(str(b_row[booking_price_col]).replace(',', '').strip())
                else:
                    booking_price = float(str(b_row.iloc[8]).replace(',', '').strip())
                
                booking_price_adjusted = round(booking_price * 0.82)
                
                print(f"    부킹원가={booking_price}, 조정가격(×0.82)={booking_price_adjusted}, 비교={round(use_price)}")
                
                if round(use_price) == booking_price_adjusted:
                    price_match = True
                    matched_booking_refs[ota_no] = matched_booking_refs.get(ota_no, 0) + 1
                    print(f"    [OK] 개별 행 매칭 성공!")
                    break
                else:
                    if log_info is None:
                        booking_file_name = booking_files[0] if booking_files else '부킹파일'
                        log_info = [name, ws_row, use_price, booking_file_name, b_idx+2, str(booking_price_adjusted), str(booking_price)]
            except Exception as e:
                print(f"    오류: {e}")
                continue
        
        if price_match:
            for cell in ws[ws_row]:
                cell.fill = fill_yellow
            print(f"    → 노란색 표시")
        else:
            if matched_booking_refs.get(ota_no, 0) > 0:
                matched_booking_refs[ota_no] -= 1
                print(f"    → 이미 매칭됨 (표시 없음)")
                continue
            
            for cell in ws[ws_row]:
                cell.font = font_red
            print(f"    [ERROR] 불일치 - 빨간색 표시 + 비교로그 기록")
            
            if log_info:
                log_ws.append(log_info)
            else:
                log_ws.append([name, ws_row, use_price, '-', '-', '불일치', '-'])

print(f"\n[완료] 부킹닷컴 비교 완료")
print("="*80)

# 익스피디아 비교 처리
print("\n" + "="*80)
print("익스피디아 비교 시작")
print("="*80)

matched_expedia_refs = {}
expedia_matched_count = 0
expedia_notfound_count = 0
expedia_mismatch_count = 0

for idx, row in df_all.iterrows():
    vendor = str(row.get('거래처', '')).strip()
    if vendor != '익스피디아':
        continue
    
    ws_row = idx + 2
    name = str(row.get(col_name_all, '')).strip()
    ota_no = str(row.get(col_ota_no, '')).strip()
    price1 = str(row.get(col_price_all_1, '')).replace(',', '').strip()
    price2 = str(row.get(col_price_all_2, '')).replace(',', '').strip()
    
    try:
        price1_f = float(price1)
    except:
        price1_f = None
    try:
        price2_f = float(price2)
    except:
        price2_f = None
    
    use_price = price2_f if price2_f else price1_f
    if use_price is None:
        continue
    
    if df_expedia.empty:
        continue
    
    # 익스피디아 CSV에서 예약번호 검색
    expedia_match = df_expedia[df_expedia[df_expedia.columns[0]] == ota_no]
    
    print(f"  행 {ws_row}: OTA번호={ota_no}, 가격={use_price}, 익스피디아매칭={len(expedia_match)}건")
    
    if expedia_match.empty:
        # 익스피디아 CSV에 예약번호 없음 -> 파란색
        print(f"    → 익스피디아 데이터에 예약번호 없음 (파란색 표시)")
        for cell in ws[ws_row]:
            cell.fill = fill_blue
        # 비교로그에 기록
        log_ws.append([name, ws_row, use_price, '익스피디아 데이터 없음', '', '', ''])
        expedia_notfound_count += 1
        continue
    
    price_match = False
    log_info = None
    
    for e_idx, e_row in expedia_match.iterrows():
        try:
            if expedia_price_col and expedia_price_col in df_expedia.columns:
                price_str = str(e_row[expedia_price_col])
            else:
                price_str = str(e_row.iloc[5])
            # "KRW 538739" 형식에서 숫자만 추출
            price_str = re.sub(r'[^\d.]', '', price_str).strip()
            expedia_price = float(price_str)
            
            # 오차 범위 1,000원 이내 허용
            price_diff = abs(use_price - expedia_price)
            
            print(f"    익스피디아가격={expedia_price}, 전체고객목록가격={use_price}, 차이={price_diff}")
            
            if price_diff <= 1000:
                price_match = True
                matched_expedia_refs[ota_no] = matched_expedia_refs.get(ota_no, 0) + 1
                print(f"    [OK] 매칭 성공! (오차 {price_diff}원)")
                break
            else:
                if log_info is None:
                    expedia_file_name = expedia_files[0] if expedia_files else '익스피디아파일'
                    log_info = [name, ws_row, use_price, expedia_file_name, e_idx+2, str(expedia_price), str(expedia_price)]
        except Exception as e:
            print(f"    오류: {e}")
            continue
    
    if price_match:
        for cell in ws[ws_row]:
            cell.fill = fill_yellow
            cell.font = Font()  # 글씨 색상 초기화 (검정색)
        print(f"    → 노란색 표시")
        expedia_matched_count += 1
    else:
        if matched_expedia_refs.get(ota_no, 0) > 0:
            matched_expedia_refs[ota_no] -= 1
            print(f"    → 이미 매칭됨 (표시 없음)")
            continue
        
        # 기존 배경색 제거 후 빨간색 글씨만 표시
        for cell in ws[ws_row]:
            cell.fill = PatternFill(fill_type=None)  # 배경색 초기화
            cell.font = font_red
        print(f"    [ERROR] 불일치 - 빨간색 표시 + 비교로그 기록")
        expedia_mismatch_count += 1
        
        if log_info:
            log_ws.append(log_info)
        else:
            log_ws.append([name, ws_row, use_price, '-', '-', '불일치', '-'])

print(f"\n[완료] 익스피디아 비교 완료")
print(f"  ✅ 매칭 성공: {expedia_matched_count}건")
print(f"  ❌ 가격 불일치: {expedia_mismatch_count}건")
print(f"  🔵 예약번호 없음: {expedia_notfound_count}건")
print("="*80)

# 결과 저장
result_path = os.path.join(dir_base, '매출_검토_결과.xlsx')
wb.save(result_path)
print(f'완료: {result_path}에 저장됨')

# Peter Ludwig 비교로그 출력
print_peter_ludwig_log()
